from __future__ import annotations
 
import inspect
import io
from html import escape
from typing import Any, Dict, List, Optional, Sequence, Tuple
 
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook as openpyxl_load_workbook
 
# =============================================================================
# 1. APPLICATION CONFIGURATION  (management assumptions only - no business data)
# =============================================================================
 
APP_TITLE = "Sales Performance Scenario Planner"
APP_SUBTITLE = (
    "Current state → selected management scenario → required future state → revenue impact"
)
 
# --- Financial-year timeline -------------------------------------------------
MONTHS_COMPLETED = 3        # April, May, June are complete
MONTHS_REMAINING = 9        # July .. March
MONTHS_JUL_JAN = 7          # July .. January
MONTHS_FEB_MAR = 2          # February, March
 
FUTURE_MONTHS: List[str] = [
    "July", "August", "September", "October",
    "November", "December", "January", "February", "March",
]
MOMENTUM_MONTHS = FUTURE_MONTHS[:MONTHS_JUL_JAN]      # July .. January
LEAKAGE_MONTHS = FUTURE_MONTHS[MONTHS_JUL_JAN:]       # February, March
MONTH_DATES = pd.date_range("2026-07-01", periods=len(FUTURE_MONTHS), freq="MS")
 
# --- Revenue assumptions (basis points) --------------------------------------
REVENUE_BPS: Dict[str, float] = {"Equity": 60.0, "Debt": 20.0, "Liquid": 10.0}
REVENUE_RATE: Dict[str, float] = {k: v / 10000.0 for k, v in REVENUE_BPS.items()}
 
ASSETS: List[str] = ["Equity", "Debt", "Liquid"]
SALES_TYPES: List[str] = ["GS", "NS"]
SALES_LABEL: Dict[str, str] = {"GS": "Gross Sales", "NS": "Net Sales"}
VERTICALS: List[str] = ["Retail", "DHNI", "VRM"]
 
# --- Scenario assumptions -----------------------------------------------------
S1_RUNRATE_UPLIFT = 0.20            # +20% on current run rate
S2_EQUITY_TARGET = 1.00             # 100% of Equity FY target by January
S2_OVERALL_TARGET = 0.75            # 75% of overall FY target by January
S3_TARGET = 1.00                    # 100% of FY target by January
S3_DEFAULT_DIP = 0.20               # default Feb-Mar run-rate dip
S4_TARGET = 1.20                    # 120% of FY target by March
S5_EQUITY_TARGET = 1.20             # 120% Equity by March
S5_OVERALL_TARGET = 1.00            # 100% overall by March
S6_SEGMENT_TARGETS: Dict[str, float] = {
    "Digital": 1.40,
    "Retail B30": 1.25,
    "Others": 1.15,
}
S7_DEFAULT_JAN_TARGET = 1.00        # January milestone = 100% of FY target
S7_DEFAULT_MAR_TARGET = 1.00        # March outcome = 100% of FY target
S7_DEFAULT_LEAKAGE = 0.20           # Feb-Mar AUM leakage / run-rate pressure
 
SEGMENT_ORDER: List[str] = ["Digital", "Retail B30", "Others"]
 
# --- Scenario navigator definitions -------------------------------------------
SCENARIOS: Dict[int, Dict[str, str]] = {
    1: {
        "label": "Scenario 1 · +20% Run-Rate Push",
        "name": "+20% Run-Rate Push",
        "kind": "runrate",
        "explanation": (
            "Increase the current Apr-Jun monthly run rate by 20% from July onward "
            "and measure the resulting March achievement."
        ),
        "milestone": "March 2027 · run rate lifted 20% for the remaining 9 months",
    },
    2: {
        "label": "Scenario 2 · 75% Overall by Jan + 100% Equity",
        "name": "75% Overall by January + 100% Equity",
        "kind": "jan_target",
        "explanation": (
            "Reach 100% of the Equity FY target and 75% of the overall FY target by January. "
            "The residual requirement is allocated to Debt and Liquid in FY-target proportion."
        ),
        "milestone": "January 2027 · Equity 100% of FY target, portfolio 75% of FY target",
    },
    3: {
        "label": "Scenario 3 · 100% by Jan, Then Feb-Mar Dip",
        "name": "100% by January, then Feb-Mar dip",
        "kind": "jan_target",
        "explanation": (
            "Reach 100% of the FY target by January, followed by a configurable "
            "February-March run-rate decline."
        ),
        "milestone": "January 2027 · 100% of FY target, then a reduced Feb-Mar run rate",
    },
    4: {
        "label": "Scenario 4 · 120% by March",
        "name": "120% by March",
        "kind": "march_target",
        "explanation": (
            "Determine the monthly run rate required to finish March at 120% of the FY target."
        ),
        "milestone": "March 2027 · 120% of FY target",
    },
    5: {
        "label": "Scenario 5 · 120% Equity + 100% Overall",
        "name": "120% Equity + 100% Overall by March",
        "kind": "march_target",
        "explanation": (
            "Reach 120% of the Equity FY target and 100% of the overall FY target by March, "
            "with Debt and Liquid balancing the remaining requirement."
        ),
        "milestone": "March 2027 · Equity 120% of FY target, portfolio 100% of FY target",
    },
    6: {
        "label": "Scenario 6 · Digital 140% + B30 125% + Others 115%",
        "name": "Digital 140% + Retail B30 125% + Others 115%",
        "kind": "march_target",
        "explanation": (
            "Model differentiated performance where Digital achieves 140%, Retail B30 achieves "
            "125% and Others achieve 115% of their respective FY targets."
        ),
        "milestone": "March 2027 · differentiated achievement by business segment",
    },
    7: {
        "label": "Scenario 7 · Momentum Build-Up to March 2027",
        "name": "Momentum Build-Up to March 2027",
        "kind": "momentum",
        "explanation": (
            "Build progressive month-on-month momentum from July 2026 to reach the January 2027 "
            "milestone, create sufficient buffer to absorb Feb-Mar AUM leakage, and protect the "
            "March 2027 target."
        ),
        "milestone": "January 2027 milestone → Feb-Mar leakage absorbed → March 2027 target held",
    },
    8: {
        "label": "Scenario 8 · Channel Growth & Target Simulator",
        "name": "Channel Growth & Target Simulator",
        "kind": "channel_simulator",
        "explanation": (
            "Independently adjust monthly growth, January 2027 target achievement and March 2027 "
            "target achievement for Digital, VRM, EM, B30, T30, T8, DHNI, Retail and Institutional."
        ),
        "milestone": "January 2027 target → February/March leakage → March 2027 target",
    },
    9: {
        "label": "Scenario 9 · Channel Mix Optimiser",
        "name": "Channel Mix Optimiser",
        "kind": "channel_optimizer",
        "explanation": (
            "Find the minimum channel growth trajectory required to achieve a selected portfolio March "
            "ambition, while preserving the January milestone and leakage assumption."
        ),
        "milestone": "Portfolio March ambition optimised across nine channels",
    },
}
SCENARIO_ORDER: List[int] = [1, 2, 3, 4, 5, 6, 7, 8, 9]
 
# Scenario 8/9 channel planning universe.  Each channel has independent
# momentum, January 2027 target and March 2027 target controls.
CHANNELS: List[str] = [
    "Digital", "VRM", "EM", "B30", "T30", "T8", "DHNI", "Retail", "Institutional"
]
S8_DEFAULT_GROWTH: Dict[str, float] = {c: 0.05 for c in CHANNELS}
S8_DEFAULT_JAN_TARGET: Dict[str, float] = {c: 1.00 for c in CHANNELS}
S8_DEFAULT_MAR_TARGET: Dict[str, float] = {c: 1.00 for c in CHANNELS}
S8_DEFAULT_LEAKAGE = 0.20
 
# --- Workbook contract --------------------------------------------------------
SHEET_ALIASES: Dict[str, List[str]] = {
    "Retail": ["RM Retail Sales", "RM Retail", "Retail Sales"],
    "DHNI": ["RM DHNI", "DHNI", "RM D-HNI"],
    "VRM": ["VRM", "RM VRM", "VRM Sales"],
}
 
# Presentation/dashboard sheet. It is deliberately NOT included in
# SHEET_ALIASES because it is not an employee-level calculation sheet.
FINAL_SHEET_ALIASES: List[str] = ["FINAL", "Final", "Final Dashboard"]
 
COLUMN_SPEC: Dict[Tuple[str, str], Dict[str, List[str]]] = {
    ("GS", "Equity"): {
        "fy": ["FY 26 TGT EQ", "Equity GS Targets"],
        "ytd_tgt": ["YTD June EQ TGT", "Q1 Equity GS Targets"],
        "ach": ["Equity GS Ach YTD June", "Equity GS Actuals"],
    },
    ("GS", "Debt"): {
        "fy": ["FY 26 TGT DT", "Debt GS Targets"],
        "ytd_tgt": ["YTD June DT TGT", "Q1 Debt GS Targets"],
        "ach": ["Debt GS Ach", "Debt GS Actuals"],
    },
    ("GS", "Liquid"): {
        "fy": ["FY 26 TGT LIQ", "Liquid GS Targets"],
        "ytd_tgt": ["YTD June LIQ TGT", "Q1 Liquid GS Targets"],
        "ach": ["Liquid GS Ach", "Liquid GS Actuals"],
    },
    ("NS", "Equity"): {
        "fy": ["FY 26 TGT EQ NS", "Equity Net Targets"],
        "ytd_tgt": ["YTD June EQ NS TGT", "Q1 Equity Net Targets"],
        "ach": ["Equity NS Ach YTD June", "Equity Net Actuals"],
    },
    ("NS", "Debt"): {
        "fy": ["FY 26 TGT DT NS", "Debt Net Targets"],
        "ytd_tgt": ["YTD June DT NS TGT", "Q1 Debt Net Targets"],
        "ach": ["Debt NS Ach", "Debt Net Actuals"],
    },
    ("NS", "Liquid"): {
        "fy": ["FY 26 TGT LIQ NS", "Liquid Net Targets"],
        "ytd_tgt": ["YTD June LIQ NS TGT", "Q1 Liquid Net Targets"],
        "ach": ["Liquid NS Ach", "Liquid Net Actuals"],
    },
}

META_ALIASES: Dict[str, List[str]] = {
    "Employee Name": ["Employee Name", "Emp Name", "Name"],
    "Emp Code": ["Emp Code", "Employee Code"],
    "ADID": ["ADID", "AD ID"],
    "Status": ["Status", "Employee Status"],
    "Type": ["Type", "Employment Type", "Functional Designation"],
    "ZONE": ["ZONE", "Zone"],
    "REGION": ["REGION", "Region"],
    "EM City": ["EM City", "City", "Location"],
    "MKT TYPE": ["MKT TYPE", "Market Type", "Mkt Type"],
}
META_FIELDS: List[str] = list(META_ALIASES.keys())
 
# -----------------------------------------------------------------------------
# SEGMENT CLASSIFICATION CONFIGURATION (Scenario 6)
# Edit this block to change how business segments are identified. The scenario
# calculation engine reads the resulting mapping and never needs to change.
# -----------------------------------------------------------------------------
SEGMENT_RULES: Dict[str, Dict[str, Any]] = {
    "Digital": {
        "search_columns": ["MKT TYPE", "Type", "REGION", "ZONE", "EM City", "Status"],
        "keywords": ["digital", "online", "d2c", "e-com", "ecom", "virtual", "vrm",
                     "inside sales", "web"],
    },
    "Retail B30": {
        "search_columns": ["MKT TYPE", "REGION", "Type"],
        "keywords": ["b30"],
    },
}
FALLBACK_SEGMENT = "Others"
 
 
# =============================================================================
# 2. GENERIC HELPERS
# =============================================================================
 
class WorkbookError(Exception):
    """Raised when the uploaded workbook does not satisfy the data contract."""
 
 
def normalize_column_name(column: Any) -> str:
    """Collapse non-breaking spaces, repeated spaces and stray padding."""
    return " ".join(str(column).replace("\u00a0", " ").strip().split())
 
 
def _norm_key(column: Any) -> str:
    return normalize_column_name(column).casefold()
 
 
def _squash_key(column: Any) -> str:
    return "".join(ch for ch in _norm_key(column) if ch.isalnum())
 
 
def normalize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of the frame with normalised column labels."""
    out = frame.copy()
    out.columns = [normalize_column_name(c) for c in out.columns]
    return out
 
 
def clean_numeric(series: Optional[pd.Series]) -> pd.Series:
    """Coerce a column to numeric, tolerating text-formatted numbers."""
    if series is None:
        return pd.Series(dtype="float64")
    if series.dtype == object:
        series = (
            series.astype(str)
            .str.replace("\u00a0", " ", regex=False)
            .str.replace(",", "", regex=False)
            .str.replace("₹", "", regex=False)
            .str.replace("%", "", regex=False)
            .str.strip()
            .replace({"": np.nan, "-": np.nan, "nan": np.nan, "NA": np.nan, "N/A": np.nan})
        )
    return pd.to_numeric(series, errors="coerce")
 
 
def as_text(series: pd.Series) -> pd.Series:
    """Coerce any column to clean text, mapping every missing marker to ''."""
    filled = series.where(series.notna(), "")
    return (
        filled.astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
        .replace({"nan": "", "NaN": "", "NaT": "", "None": "", "<NA>": ""})
    )
 
 
def text_column(frame: pd.DataFrame, column: str) -> pd.Series:
    """Safe accessor for a metadata column that may be absent."""
    if column not in frame.columns:
        return pd.Series([""] * len(frame), index=frame.index, dtype=object)
    return as_text(frame[column])
 
 
def safe_div(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    """Division that never raises and never returns inf/nan."""
    try:
        if numerator is None or denominator is None:
            return None
        n = float(numerator)
        d = float(denominator)
        if not np.isfinite(n) or not np.isfinite(d) or abs(d) < 1e-12:
            return None
        result = n / d
        return result if np.isfinite(result) else None
    except (TypeError, ValueError):
        return None
 
 
def _num(value: Any) -> Optional[float]:
    """Normalise a possibly-missing numeric to float or None."""
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    return f if np.isfinite(f) else None
 
 
def _z(value: Any) -> float:
    """Numeric value with missing treated as zero (for summation)."""
    v = _num(value)
    return 0.0 if v is None else v
 
 
def _ssum(series: pd.Series) -> Optional[float]:
    """Sum that returns None when every entry is missing."""
    total = series.sum(min_count=1)
    return _num(total)
 
 
# --- Display formatting -------------------------------------------------------
 
NA_TEXT = "0"
 
 
def fmt_cr(value: Any, decimals: int = 0) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"₹ {v:,.{decimals}f} Cr"
 
 
def fmt_cr_signed(value: Any, decimals: int = 0) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"₹ {v:+,.{decimals}f} Cr"
 
 
def fmt_pct(value: Any, decimals: int = 1) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v * 100:,.{decimals}f}%"
 
 
def fmt_pct_signed(value: Any, decimals: int = 1) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v * 100:+,.{decimals}f}%"
 
 
def fmt_pts(value: Any, decimals: int = 1) -> str:
    """Percentage-point delta, e.g. +18.4 pts."""
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v * 100:+,.{decimals}f} pts"
 
 
def fmt_num(value: Any, decimals: int = 0) -> str:
    v = _num(value)
    if v is None:
        return NA_TEXT
    return f"{v:,.{decimals}f}"
 
 
FORMATTERS = {
    "cr": fmt_cr,
    "cr1": lambda v: fmt_cr(v, 1),
    "cr_signed": fmt_cr_signed,
    "cr1_signed": lambda v: fmt_cr_signed(v, 1),
    "pct": fmt_pct,
    "pct_signed": fmt_pct_signed,
    "pts": fmt_pts,
    "num": fmt_num,
    "txt": lambda v: NA_TEXT if v is None or (isinstance(v, float) and not np.isfinite(v)) else str(v),
}
 
 
def format_table(frame: pd.DataFrame, formats: Dict[str, str]) -> pd.DataFrame:
    """Return a display-ready copy of a numeric frame using the given formats."""
    out = pd.DataFrame(index=frame.index)
    for column in frame.columns:
        kind = formats.get(column, "txt")
        formatter = FORMATTERS.get(kind, FORMATTERS["txt"])
        out[column] = [formatter(v) for v in frame[column]]
    return out
 
 
# =============================================================================
# 3. WORKBOOK LOADING, VALIDATION & CLEANING
# =============================================================================
 
def _build_column_index(frame: pd.DataFrame) -> Dict[str, int]:
    """Map normalised column keys to positional index (first occurrence wins)."""
    index: Dict[str, int] = {}
    for position, column in enumerate(frame.columns):
        for key in (_norm_key(column), _squash_key(column)):
            if key and key not in index:
                index[key] = position
    return index
 
 
def _resolve_column(index: Dict[str, int], aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        for key in (_norm_key(alias), _squash_key(alias)):
            if key in index:
                return index[key]
    return None
 
 
def _expected_header_keys() -> set:
    keys = set()
    for spec in COLUMN_SPEC.values():
        for aliases in spec.values():
            for alias in aliases:
                keys.add(_norm_key(alias))
    for aliases in META_ALIASES.values():
        for alias in aliases:
            keys.add(_norm_key(alias))
    return keys
 
 
def _detect_header_row(raw: pd.DataFrame, max_scan: int = 15) -> int:
    """Find the row that actually holds the column headers."""
    expected = _expected_header_keys()
    best_row, best_score = 0, -1
    for row in range(min(max_scan, len(raw))):
        values = {_norm_key(v) for v in raw.iloc[row].tolist() if str(v).strip().lower() != "nan"}
        score = len(values & expected)
        if score > best_score:
            best_row, best_score = row, score
    return best_row if best_score >= 4 else 0
 
 
def _match_sheet(available: Sequence[str], aliases: Sequence[str]) -> Optional[str]:
    normalised = {_norm_key(s): s for s in available}
    for alias in aliases:
        key = _norm_key(alias)
        if key in normalised:
            return normalised[key]
    for alias in aliases:
        key = _norm_key(alias)
        for sheet_key, sheet in normalised.items():
            if key in sheet_key:
                return sheet
    return None
 
 
def validate_frame(frame: pd.DataFrame, index: Dict[str, int], sheet_label: str) -> List[str]:
    """Return a list of human-readable descriptions of missing required columns."""
    missing: List[str] = []
    for (sales, asset), spec in COLUMN_SPEC.items():
        for role, aliases in spec.items():
            if _resolve_column(index, aliases) is None:
                role_label = {"fy": "FY target", "ytd_tgt": "YTD June target", "ach": "YTD June achievement"}[role]
                missing.append(
                    f"{sheet_label} · {SALES_LABEL[sales]} · {asset} {role_label} "
                    f"(expected column '{aliases[0]}')"
                )
    return missing
 
 
def _extract_records(frame: pd.DataFrame, vertical: str) -> pd.DataFrame:
    """Turn one workbook sheet into a tidy per-employee record frame."""
    index = _build_column_index(frame)
    records = pd.DataFrame(index=frame.index)
    records["Vertical"] = vertical
 
    for field, aliases in META_ALIASES.items():
        position = _resolve_column(index, aliases)
        if position is None:
            records[field] = ""
        else:
            records[field] = as_text(frame.iloc[:, position]).to_numpy()
 
    for (sales, asset), spec in COLUMN_SPEC.items():
        for role, aliases in spec.items():
            position = _resolve_column(index, aliases)
            series = frame.iloc[:, position] if position is not None else None
            records[f"{sales}_{asset}_{role}"] = clean_numeric(series).to_numpy()
 
    return records
 
 
def _clean_records(records: pd.DataFrame) -> pd.DataFrame:
    """Drop non-employee rows while preserving legitimate negative values."""
    names = text_column(records, "Employee Name").str.casefold()
    invalid = names.isin({"", "nan", "none", "total", "grand total", "sum", "subtotal"})
    numeric_columns = [c for c in records.columns if c.split("_")[0] in SALES_TYPES]
    empty_rows = records[numeric_columns].isna().all(axis=1)
    cleaned = records.loc[~(invalid | empty_rows)].copy()
    cleaned[numeric_columns] = cleaned[numeric_columns].fillna(0.0)
    return cleaned.reset_index(drop=True)
 
 
@st.cache_data(show_spinner=False)
def load_workbook(payload: bytes) -> pd.DataFrame:
    """Read, validate and clean the workbook into a single tidy record frame."""
    try:
        excel = pd.ExcelFile(io.BytesIO(payload), engine="openpyxl")
    except Exception as exc:  # pragma: no cover - defensive
        raise WorkbookError(
            "The file could not be opened as an Excel workbook. "
            "Please upload a valid .xlsx file."
        ) from exc
 
    available = list(excel.sheet_names)
    resolved: Dict[str, str] = {}
    for vertical, aliases in SHEET_ALIASES.items():
        sheet = _match_sheet(available, aliases)
        if sheet is not None:
            resolved[vertical] = sheet
 
    missing_sheets = [v for v in SHEET_ALIASES if v not in resolved]
    if missing_sheets:
        wanted = ", ".join(f"'{SHEET_ALIASES[v][0]}'" for v in missing_sheets)
        raise WorkbookError(
            f"The workbook is missing the required calculation sheet(s): {wanted}. "
            "Please upload the standard RM scorecard workbook."
        )
 
    frames: List[pd.DataFrame] = []
    problems: List[str] = []
    for vertical, sheet in resolved.items():
        raw = pd.read_excel(excel, sheet_name=sheet, header=None, nrows=20)
        header_row = _detect_header_row(raw)
        frame = normalize_frame(pd.read_excel(excel, sheet_name=sheet, header=header_row))
        index = _build_column_index(frame)
        problems.extend(validate_frame(frame, index, vertical))
        if not problems:
            frames.append(_extract_records(frame, vertical))
 
    if problems:
        raise WorkbookError(
            "The workbook is missing required columns:\n\n- " + "\n- ".join(problems[:12])
            + ("\n- …" if len(problems) > 12 else "")
        )
 
    records = _clean_records(pd.concat(frames, ignore_index=True))
    if records.empty:
        raise WorkbookError("No employee records were found in the calculation sheets.")
    return records
 
 
 
# =============================================================================
# 3A. FINAL SHEET — MANAGEMENT DASHBOARD VIEW
# =============================================================================
 
def _find_final_sheet_name(sheet_names: Sequence[str]) -> Optional[str]:
    """Resolve the sixth workbook sheet named FINAL."""
    return _match_sheet(sheet_names, FINAL_SHEET_ALIASES)
 
 
def _excel_rgb(color: Any) -> Optional[str]:
    """Convert an openpyxl RGB colour to a CSS hex colour when possible."""
    try:
        if color is None or color.type != "rgb" or not color.rgb:
            return None
        raw = str(color.rgb)
        rgb = raw[-6:]
        if len(rgb) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in rgb):
            return f"#{rgb}"
    except Exception:
        return None
    return None
 
 
def _display_excel_value(value: Any, number_format: str = "") -> str:
    """Format Excel values for the FINAL dashboard without exposing formulas."""
    if value is None:
        return ""
 
    if isinstance(value, (pd.Timestamp,)):
        return value.strftime("%d-%b-%Y")
 
    # Date/datetime objects from openpyxl.
    if hasattr(value, "strftime") and not isinstance(value, str):
        try:
            return value.strftime("%d-%b-%Y")
        except Exception:
            pass
 
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        numeric = float(value)
        if not np.isfinite(numeric):
            return ""
 
        fmt = str(number_format or "")
        if "%" in fmt:
            # Excel stores percentages as fractions.
            decimals = 0
            if "." in fmt.split("%")[0]:
                decimals = len(fmt.split("%")[0].split(".")[-1].replace("0", "0"))
            decimals = min(max(decimals, 0), 2)
            return f"{numeric * 100:,.{decimals}f}%"
 
        # Respect accounting-style parentheses when possible.
        negative_parentheses = numeric < 0 and "(" in fmt and ")" in fmt
        abs_value = abs(numeric)
 
        if abs(abs_value - round(abs_value)) < 1e-9:
            rendered = f"{abs_value:,.0f}"
        else:
            rendered = f"{abs_value:,.2f}".rstrip("0").rstrip(".")
 
        if numeric < 0:
            return f"({rendered})" if negative_parentheses else f"-{rendered}"
        return rendered
 
    return str(value)
 
 
@st.cache_data(show_spinner=False)
def load_final_sheet_frame(payload: bytes) -> pd.DataFrame:
    """
    Read the FINAL sheet as raw cells for a fallback / inspection table.
 
    The sheet intentionally has no single header row, so header=None is used.
    """
    excel = pd.ExcelFile(io.BytesIO(payload), engine="openpyxl")
    sheet = _find_final_sheet_name(excel.sheet_names)
    if sheet is None:
        raise WorkbookError(
            "The workbook does not contain the sixth sheet 'FINAL'. "
            "Please upload the workbook that contains Summary, Summary-Achievement, "
            "RM Retail Sales, RM DHNI, VRM and FINAL."
        )
 
    frame = pd.read_excel(excel, sheet_name=sheet, header=None)
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return frame.reset_index(drop=True)
 
 
@st.cache_data(show_spinner=False)
def build_final_sheet_html(payload: bytes) -> str:
    """
    Render the Excel FINAL sheet into a scrollable HTML table.
 
    This keeps merged headings and the workbook's basic font/fill/alignment
    styling, while using cached formula results (data_only=True).
    """
    workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    sheet_name = _find_final_sheet_name(workbook.sheetnames)
    if sheet_name is None:
        raise WorkbookError(
            "The workbook does not contain the sixth sheet 'FINAL'. "
            "Please upload the workbook that contains the FINAL dashboard sheet."
        )
 
    ws = workbook[sheet_name]
 
    # Find the real content boundary from non-empty values rather than Excel's
    # formatted used-range, which can extend thousands of blank rows/columns.
    non_empty = [
        (cell.row, cell.column)
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    ]
    if not non_empty:
        return "<div class='note'>The FINAL sheet is empty.</div>"
 
    min_row = min(r for r, _ in non_empty)
    max_row = max(r for r, _ in non_empty)
    min_col = min(c for _, c in non_empty)
    max_col = max(c for _, c in non_empty)
 
    # Merged-cell lookup.
    merge_anchor: Dict[Tuple[int, int], Tuple[int, int]] = {}
    merge_covered: set = set()
    for merged in ws.merged_cells.ranges:
        # Only consider merged ranges intersecting the content boundary.
        if (
            merged.max_row < min_row or merged.min_row > max_row
            or merged.max_col < min_col or merged.min_col > max_col
        ):
            continue
        anchor = (merged.min_row, merged.min_col)
        merge_anchor[anchor] = (
            merged.max_row - merged.min_row + 1,
            merged.max_col - merged.min_col + 1,
        )
        for rr in range(merged.min_row, merged.max_row + 1):
            for cc in range(merged.min_col, merged.max_col + 1):
                if (rr, cc) != anchor:
                    merge_covered.add((rr, cc))
 
    html_parts = [
        """
        <div style="
            overflow-x:auto;
            overflow-y:auto;
            max-height:78vh;
            border:1px solid rgba(76,99,133,.22);
            border-radius:12px;
            background:rgba(255,255,255,.72);
            padding:4px;
        ">
        <table style="
            border-collapse:collapse;
            width:max-content;
            min-width:100%;
            font-family:Arial, sans-serif;
            font-size:12px;
            color:#172033;
        ">
        """
    ]
 
    for row_idx in range(min_row, max_row + 1):
        row_values = [
            ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(min_col, max_col + 1)
        ]
 
        # Keep dashboard spacing, but collapse very large blank areas to a
        # single slim spacer row.
        if all(v in (None, "") for v in row_values):
            html_parts.append(
                "<tr><td colspan='{}' style='height:8px;border:none;background:rgba(255,255,255,.72);'></td></tr>"
                .format(max_col - min_col + 1)
            )
            continue
 
        html_parts.append("<tr>")
        for col_idx in range(min_col, max_col + 1):
            if (row_idx, col_idx) in merge_covered:
                continue
 
            cell = ws.cell(row=row_idx, column=col_idx)
            rowspan, colspan = merge_anchor.get((row_idx, col_idx), (1, 1))
 
            value = _display_excel_value(cell.value, cell.number_format)
            font_color = _excel_rgb(cell.font.color)
            fill_color = _excel_rgb(cell.fill.fgColor)
 
            styles = [
                "border:1px solid rgba(76,99,133,.18)",
                "padding:5px 7px",
                "min-width:78px",
                "white-space:nowrap",
                "vertical-align:middle",
                "background:rgba(255,255,255,.72)",
            ]
 
            if fill_color and fill_color.lower() not in {"#000000", "#ffffff"}:
                styles.append(f"background:{fill_color}")
            if font_color:
                styles.append(f"color:{font_color}")
            if cell.font.bold:
                styles.append("font-weight:700")
            if cell.font.italic:
                styles.append("font-style:italic")
 
            horizontal = getattr(cell.alignment, "horizontal", None)
            if horizontal in {"center", "centerContinuous"}:
                styles.append("text-align:center")
            elif horizontal == "right":
                styles.append("text-align:right")
            else:
                styles.append("text-align:left")
 
            attrs = []
            if rowspan > 1:
                attrs.append(f"rowspan='{rowspan}'")
            if colspan > 1:
                attrs.append(f"colspan='{colspan}'")
 
            html_parts.append(
                f"<td {' '.join(attrs)} style=\"{';'.join(styles)}\">"
                f"{escape(value)}"
                "</td>"
            )
        html_parts.append("</tr>")
 
    html_parts.append("</table></div>")
    return "".join(html_parts)
 
 
FINAL_METRIC_ROWS: List[str] = [
    "Overall", "Equity", "Debt", "Liquid",
    "Retail", "DHNI", "VRM", "Insti", "Digital",
    "Alternatives", "Passives",
]
 
 
def _final_key(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").strip().split()).lower()
 
 
def _final_number(value: Any) -> Optional[float]:
    """Convert FINAL-sheet numeric / accounting text into float."""
    if value is None:
        return None
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        numeric = float(value)
        return numeric if np.isfinite(numeric) else None
 
    raw = str(value).strip()
    if not raw or raw.lower() in {"-", "—", "na", "n/a", "none", "nan", "#div/0!"}:
        return None
 
    negative = raw.startswith("(") and raw.endswith(")")
    cleaned = (
        raw.replace(",", "")
        .replace("₹", "")
        .replace("%", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )
    try:
        numeric = float(cleaned)
    except ValueError:
        return None
    if negative:
        numeric = -numeric
    return numeric
 
 
def _final_known_label(value: Any) -> Optional[str]:
    key = _final_key(value)
    aliases = {
        "overall": "Overall",
        "equity": "Equity",
        "debt": "Debt",
        "liquid": "Liquid",
        "retail": "Retail",
        "dhni": "DHNI",
        "vrm": "VRM",
        "insti": "Insti",
        "institutional": "Insti",
        "digital": "Digital",
        "alternatives": "Alternatives",
        "alternate": "Alternatives",
        "passives": "Passives",
        "passive": "Passives",
    }
    return aliases.get(key)
 
 
def _scan_final_sheet(ws: Any) -> Tuple[int, int]:
    """Cap the scan to the management-dashboard area, not formatted blank Excel space."""
    return min(max(ws.max_row, 1), 320), min(max(ws.max_column, 1), 180)
 
 
def _find_final_cells(ws: Any, wanted: str) -> List[Tuple[int, int]]:
    wanted_key = _final_key(wanted)
    max_row, max_col = _scan_final_sheet(ws)
    found: List[Tuple[int, int]] = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if _final_key(ws.cell(row=row, column=col).value) == wanted_key:
                found.append((row, col))
    return found
 
 
def _find_header_near(
    ws: Any,
    title_position: Tuple[int, int],
    header_name: str,
    row_window: int = 8,
    col_before: int = 4,
    col_after: int = 24,
) -> Optional[Tuple[int, int]]:
    title_row, title_col = title_position
    max_row, max_col = _scan_final_sheet(ws)
    wanted = _final_key(header_name)
    for row in range(title_row, min(title_row + row_window, max_row) + 1):
        start_col = max(1, title_col - col_before)
        end_col = min(max_col, title_col + col_after)
        for col in range(start_col, end_col + 1):
            if _final_key(ws.cell(row=row, column=col).value) == wanted:
                return row, col
    return None
 
 
def _augment_final_runrate(frame: pd.DataFrame, months_done: int) -> pd.DataFrame:
    """Recreate the run-rate formulas shown in FINAL from Target and YTD."""
    if frame.empty:
        return frame
 
    months = max(int(months_done), 1)
    out = frame.copy()
    out["FY27 Target"] = pd.to_numeric(out["FY27 Target"], errors="coerce")
    out["YTD"] = pd.to_numeric(out["YTD"], errors="coerce")
 
    out["Achievement %"] = np.where(
        (out["FY27 Target"] > 0) & (out["YTD"] >= 0),
        out["YTD"] / out["FY27 Target"],
        np.nan,
    )
    out["Current RR"] = out["YTD"] / months
 
    # This matches the FINAL workbook screenshot:
    # 154,757 / 12 = 12,896; 20,699 / 12 = 1,725.
    out["Required RR to Target"] = out["FY27 Target"] / 12.0
 
    # Annualise the current Apr-Jun run rate to a full 12-month FY.
    out["Estimated FY @ Current RR"] = out["Current RR"] * 12.0
    out["Projected FY %"] = np.where(
        (out["FY27 Target"] > 0) & (out["Estimated FY @ Current RR"] >= 0),
        out["Estimated FY @ Current RR"] / out["FY27 Target"],
        np.nan,
    )
    return out
 
 
def _parse_final_sales_block(ws: Any, title: str, months_done: int) -> pd.DataFrame:
    """Parse the NET SALES / GROSS SALES run-rate block on the FINAL sheet."""
    positions = _find_final_cells(ws, title)
    for position in positions:
        header = _find_header_near(ws, position, "FY27 Target")
        if header is None:
            continue
 
        header_row, target_col = header
 
        # In the FINAL block the row label sits immediately to the left of FY27 Target.
        label_col = max(1, target_col - 1)
        ytd_col = target_col + 1
 
        rows: List[Dict[str, Any]] = []
        seen: set = set()
        max_row, _ = _scan_final_sheet(ws)
 
        for row in range(header_row + 1, min(header_row + 28, max_row) + 1):
            label = _final_known_label(ws.cell(row=row, column=label_col).value)
            if label is None or label in seen:
                continue
 
            target = _final_number(ws.cell(row=row, column=target_col).value)
            ytd = _final_number(ws.cell(row=row, column=ytd_col).value)
 
            # Ignore title/spacer rows accidentally matching a label.
            if target is None and ytd is None:
                continue
 
            rows.append({"Metric": label, "FY27 Target": target, "YTD": ytd})
            seen.add(label)
 
        if rows:
            frame = pd.DataFrame(rows).set_index("Metric")
            order = [label for label in FINAL_METRIC_ROWS if label in frame.index]
            return _augment_final_runrate(frame.loc[order].reset_index(), months_done).set_index("Metric")
 
    return pd.DataFrame()
 
 
def _parse_final_aum_block(ws: Any) -> pd.DataFrame:
    """Parse Target / Current AUM from the FINAL management sheet."""
    positions = _find_final_cells(ws, "AUM")
    for position in positions:
        title_row, title_col = position
        max_row, max_col = _scan_final_sheet(ws)
 
        header_row = None
        target_col = None
        current_col = None
 
        for row in range(title_row, min(title_row + 6, max_row) + 1):
            for col in range(max(1, title_col - 4), min(max_col, title_col + 8) + 1):
                if _final_key(ws.cell(row=row, column=col).value) == "target":
                    # Find Current on the same header row.
                    for cc in range(col + 1, min(max_col, col + 5) + 1):
                        if _final_key(ws.cell(row=row, column=cc).value) == "current":
                            header_row = row
                            target_col = col
                            current_col = cc
                            break
                if header_row is not None:
                    break
            if header_row is not None:
                break
 
        if header_row is None or target_col is None or current_col is None:
            continue
 
        label_col = max(1, target_col - 1)
        rows: List[Dict[str, Any]] = []
        seen: set = set()
 
        for row in range(header_row + 1, min(header_row + 28, max_row) + 1):
            label = _final_known_label(ws.cell(row=row, column=label_col).value)
            if label is None or label in seen:
                continue
            target = _final_number(ws.cell(row=row, column=target_col).value)
            current = _final_number(ws.cell(row=row, column=current_col).value)
            if target is None and current is None:
                continue
            rows.append({"Metric": label, "Target": target, "Current": current})
            seen.add(label)
 
        if rows:
            frame = pd.DataFrame(rows).set_index("Metric")
            order = [label for label in FINAL_METRIC_ROWS if label in frame.index]
            frame = frame.loc[order].copy()
            frame["Achievement %"] = np.where(
                (frame["Target"] > 0) & (frame["Current"] >= 0),
                frame["Current"] / frame["Target"],
                np.nan,
            )
            frame["Gap to Target"] = frame["Target"] - frame["Current"]
            return frame
 
    return pd.DataFrame()
 
 
def _parse_months_done(ws: Any) -> int:
    positions = (
        _find_final_cells(ws, "#months done")
        + _find_final_cells(ws, "months done")
        + _find_final_cells(ws, "# months done")
    )
    max_row, max_col = _scan_final_sheet(ws)
 
    for row, col in positions:
        # Search immediately around / below the label for the red "3".
        for rr in range(row, min(row + 4, max_row) + 1):
            for cc in range(max(1, col - 2), min(max_col, col + 4) + 1):
                value = _final_number(ws.cell(row=rr, column=cc).value)
                if value is not None and 1 <= value <= 12:
                    return int(round(value))
    return MONTHS_COMPLETED
 
 
@st.cache_data(show_spinner=False)
def parse_final_dashboard_metrics(payload: bytes) -> Dict[str, Any]:
    """Return structured management metrics from the workbook's FINAL sheet."""
    workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    sheet_name = _find_final_sheet_name(workbook.sheetnames)
    if sheet_name is None:
        raise WorkbookError(
            "The workbook is missing the required sixth sheet 'FINAL'. "
            "Please upload the workbook containing FINAL."
        )
 
    ws = workbook[sheet_name]
    months_done = _parse_months_done(ws)
 
    gs = _parse_final_sales_block(ws, "GROSS SALES", months_done)
    ns = _parse_final_sales_block(ws, "NET SALES", months_done)
    aum = _parse_final_aum_block(ws)
 
    return {
        "sheet_name": sheet_name,
        "months_done": months_done,
        "GS": gs,
        "NS": ns,
        "AUM": aum,
    }
 
 
def _model_metric_baseline(model: "ScenarioModel", sales: str, label: str) -> Dict[str, Any]:
    if label == "Overall":
        return model.baseline(sales)
    if label in ASSETS:
        return model.baseline(sales, asset=label)
    if label in VERTICALS:
        return model.baseline(sales, vertical=label)
    return {}
 
 
def _model_metric_cell(model: "ScenarioModel", sales: str, label: str) -> Optional[Dict[str, Any]]:
    if label == "Overall":
        return model.cell(sales)
    if label in ASSETS:
        return model.cell(sales, asset=label)
    if label in VERTICALS:
        return model.cell(sales, vertical=label)
    return None
 
 
def final_sales_metrics(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
    sales: str,
) -> pd.DataFrame:
    """
    Use FINAL Target/YTD as the visible source of truth.
    Fill modelled Overall/asset/vertical rows from the scenario data only if
    the FINAL parser could not locate them.
    """
    parsed = final_metrics.get(sales)
    if isinstance(parsed, pd.DataFrame) and not parsed.empty:
        frame = parsed.copy()
    else:
        frame = pd.DataFrame()
 
    months_done = int(final_metrics.get("months_done", MONTHS_COMPLETED))
    needed = ["Overall", *ASSETS, *VERTICALS]
 
    fallback_rows: List[Dict[str, Any]] = []
    for label in needed:
        if not frame.empty and label in frame.index:
            continue
        base = _model_metric_baseline(model, sales, label)
        if not base:
            continue
        fallback_rows.append(
            {
                "Metric": label,
                "FY27 Target": base.get("fy_target"),
                "YTD": base.get("ytd_ach"),
            }
        )
 
    if fallback_rows:
        fallback = _augment_final_runrate(pd.DataFrame(fallback_rows), months_done).set_index("Metric")
        if frame.empty:
            frame = fallback
        else:
            frame = pd.concat([frame, fallback], axis=0)
 
    if frame.empty:
        return frame
 
    # Preserve management ordering and avoid accidental duplicate rows.
    frame = frame.loc[~frame.index.duplicated(keep="first")].copy()
    order = [label for label in FINAL_METRIC_ROWS if label in frame.index]
    return frame.loc[order]
 
 
def build_final_scenario_comparison(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
    sales: str,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario comparison expressed on the same Target/YTD/run-rate metrics as FINAL."""
    current = final_sales_metrics(final_metrics, model, sales)
    rows: List[Dict[str, Any]] = []
 
    for label in current.index.tolist():
        source = current.loc[label]
        cell = _model_metric_cell(model, sales, label)
 
        final_target = _num(source.get("FY27 Target"))
        current_rr = _num(source.get("Current RR"))
        current_projection = _num(source.get("Estimated FY @ Current RR"))
        current_pct = _num(source.get("Projected FY %"))
 
        model_base = _model_metric_baseline(model, sales, label)
        model_target = _num(model_base.get("fy_target")) if model_base else None
 
        scenario_pct = _num(cell.get("march_pct")) if cell is not None else None
 
        scenario_amount = None
        scenario_rr = None
        rr_change = None
        delta_pp = None
 
        if cell is not None:
            # Anchor the scenario outcome to the FINAL FY27 target so the
            # management comparison uses one common metric base.
            scenario_amount = (
                final_target * scenario_pct
                if final_target is not None and scenario_pct is not None
                else _num(cell.get("march_amount"))
            )
 
            scenario_rr = _num(cell.get("scen_rr"))
            if (
                scenario_rr is not None
                and final_target is not None
                and model_target is not None
                and model_target != 0
            ):
                scenario_rr = scenario_rr * final_target / model_target
 
            if current_rr is not None and scenario_rr is not None and current_rr != 0:
                rr_change = scenario_rr / current_rr - 1.0
 
            if current_pct is not None and scenario_pct is not None:
                delta_pp = scenario_pct - current_pct
 
        rows.append(
            {
                "Metric": label,
                "FY27 Target": final_target,
                "YTD": _num(source.get("YTD")),
                "Current RR": current_rr,
                "Required RR to Target": _num(source.get("Required RR to Target")),
                "Current FY Estimate": current_projection,
                "Current Projected %": current_pct,
                "Scenario / Required RR": scenario_rr,
                "Run Rate Change %": rr_change,
                "Scenario March Estimate": scenario_amount,
                "Scenario March %": scenario_pct,
                "Scenario Δ pp": delta_pp,
            }
        )
 
    formats = {
        "Metric": "txt",
        "FY27 Target": "cr",
        "YTD": "cr",
        "Current RR": "cr",
        "Required RR to Target": "cr",
        "Current FY Estimate": "cr",
        "Current Projected %": "pct",
        "Scenario / Required RR": "cr",
        "Run Rate Change %": "pct_signed",
        "Scenario March Estimate": "cr",
        "Scenario March %": "pct",
        "Scenario Δ pp": "pts",
    }
    return pd.DataFrame(rows), formats
 
 
FINAL_ASSET_ROWS: List[str] = ["Equity", "Debt", "Liquid"]
FINAL_CHANNEL_ROWS: List[str] = [
    "Retail", "DHNI", "VRM", "Insti", "Digital", "Alternatives", "Passives",
]
 
 
def _render_metric_bifurcation(
    frame: pd.DataFrame,
    formats: Dict[str, str],
    *,
    metric_column: str = "Metric",
) -> None:
    """
    Management-friendly table split:
      • Overall
      • Asset Class -> Equity / Debt / Liquid
      • Channel -> Retail / DHNI / VRM / Insti / Digital / ...
    """
    if frame is None or frame.empty or metric_column not in frame.columns:
        st.info("No metrics are available for this view.")
        return
 
    work = frame.copy()
 
    overall = work.loc[work[metric_column] == "Overall"].copy()
    assets = work.loc[work[metric_column].isin(FINAL_ASSET_ROWS)].copy()
    channels = work.loc[
        ~work[metric_column].isin(["Overall", *FINAL_ASSET_ROWS])
    ].copy()
 
    if not overall.empty:
        st.markdown("<div class='subsection-title'>Overall</div>", unsafe_allow_html=True)
        overall = overall.rename(columns={metric_column: "Scope"})
        overall["Scope"] = "Overall"
        overall_formats = dict(formats)
        overall_formats.pop(metric_column, None)
        overall_formats["Scope"] = "txt"
        show_table(overall, overall_formats)
 
    if not assets.empty:
        st.markdown("<div class='subsection-title'>Asset Class</div>", unsafe_allow_html=True)
        assets = assets.rename(columns={metric_column: "Asset Class"})
        asset_formats = dict(formats)
        asset_formats.pop(metric_column, None)
        asset_formats["Asset Class"] = "txt"
        show_table(assets, asset_formats)
 
    if not channels.empty:
        st.markdown("<div class='subsection-title'>Channel</div>", unsafe_allow_html=True)
        channels = channels.rename(columns={metric_column: "Channel"})
        channel_formats = dict(formats)
        channel_formats.pop(metric_column, None)
        channel_formats["Channel"] = "txt"
        show_table(channels, channel_formats)
 
 
def render_current_runrate_metric_grid(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
) -> None:
    """Current FINAL run-rate grid, shown after current-vs-scenario comparison."""
    section("Current Run-Rate Metric Grid")
 
    display_cols = [
        "FY27 Target",
        "YTD",
        "Achievement %",
        "Current RR",
        "Required RR to Target",
        "Estimated FY @ Current RR",
        "Projected FY %",
    ]
    formats = {
        "Metric": "txt",
        "FY27 Target": "cr",
        "YTD": "cr",
        "Achievement %": "pct",
        "Current RR": "cr",
        "Required RR to Target": "cr",
        "Estimated FY @ Current RR": "cr",
        "Projected FY %": "pct",
    }
 
    tabs = st.tabs(["Gross Sales", "Net Sales"])
    for tab, sales in zip(tabs, SALES_TYPES):
        with tab:
            frame = final_sales_metrics(final_metrics, model, sales)
            if frame.empty:
                st.info(f"{SALES_LABEL[sales]} metrics could not be located in FINAL.")
                continue
 
            display = frame.reset_index()[["Metric", *display_cols]]
            _render_metric_bifurcation(display, formats)
 
 
def render_final_metric_baseline(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
) -> None:
    """Headline management metrics sourced from FINAL."""
    section("Current Performance Metrics · FINAL")
 
    gs = final_sales_metrics(final_metrics, model, "GS")
    ns = final_sales_metrics(final_metrics, model, "NS")
 
    def overall(frame: pd.DataFrame) -> pd.Series:
        return (
            frame.loc["Overall"]
            if not frame.empty and "Overall" in frame.index
            else pd.Series(dtype=float)
        )
 
    gs_o = overall(gs)
    ns_o = overall(ns)
 
    kpi_row([
        ("GS FY27 Target", fmt_cr(gs_o.get("FY27 Target")), "FINAL", "off"),
        ("GS YTD", fmt_cr(gs_o.get("YTD")),
         fmt_pct(gs_o.get("Achievement %"))
         if _num(gs_o.get("Achievement %")) is not None else None),
        ("GS Current RR", fmt_cr(gs_o.get("Current RR")), "Current pace", "off"),
        ("GS Required RR", fmt_cr(gs_o.get("Required RR to Target")), "Target pace", "off"),
        ("GS Projected FY %", fmt_pct(gs_o.get("Projected FY %")),
         None if _num(gs_o.get("Projected FY %")) is None
         else fmt_pts(gs_o.get("Projected FY %") - 1.0)),
    ])
 
    kpi_row([
        ("NS FY27 Target", fmt_cr(ns_o.get("FY27 Target")), "FINAL", "off"),
        ("NS YTD", fmt_cr(ns_o.get("YTD")),
         fmt_pct(ns_o.get("Achievement %"))
         if _num(ns_o.get("Achievement %")) is not None else None),
        ("NS Current RR", fmt_cr(ns_o.get("Current RR")), "Current pace", "off"),
        ("NS Required RR", fmt_cr(ns_o.get("Required RR to Target")), "Target pace", "off"),
        ("NS Projected FY %", fmt_pct(ns_o.get("Projected FY %")),
         None if _num(ns_o.get("Projected FY %")) is None
         else fmt_pts(ns_o.get("Projected FY %") - 1.0)),
    ])
 
    st.markdown(
        "<div class='note'>Headline performance is sourced from the workbook's "
        "<b>FINAL</b> sheet. Current RR shows the observed YTD pace; Required RR "
        "shows the monthly pace implied by the full-year target.</div>",
        unsafe_allow_html=True,
    )
 
 
def render_final_scenario_comparison(
    final_metrics: Dict[str, Any],
    model: "ScenarioModel",
    basis: str,
) -> None:
    """Selected scenario compared with the same FINAL metrics used in the baseline."""
    section(f"Current Metrics vs Scenario {model.scenario_id} · {model.meta['name']}")
 
    gs_frame, gs_formats = build_final_scenario_comparison(final_metrics, model, "GS")
    ns_frame, ns_formats = build_final_scenario_comparison(final_metrics, model, "NS")
 
    def overall_row(frame: pd.DataFrame) -> pd.Series:
        if frame.empty:
            return pd.Series(dtype=float)
        match = frame.loc[frame["Metric"] == "Overall"]
        return match.iloc[0] if not match.empty else pd.Series(dtype=float)
 
    gs_o = overall_row(gs_frame)
    ns_o = overall_row(ns_frame)
 
    kpi_row([
        ("GS Current Projection", fmt_pct(gs_o.get("Current Projected %")),
         fmt_cr(gs_o.get("Current FY Estimate")), "off"),
        ("GS Scenario Projection", fmt_pct(gs_o.get("Scenario March %")),
         None if _num(gs_o.get("Scenario Δ pp")) is None
         else fmt_pts(gs_o.get("Scenario Δ pp"))),
        ("GS Current RR", fmt_cr(gs_o.get("Current RR")), "Current pace", "off"),
        ("GS Scenario RR", fmt_cr(gs_o.get("Scenario / Required RR")),
         fmt_pct_signed(gs_o.get("Run Rate Change %"))
         if _num(gs_o.get("Run Rate Change %")) is not None else None),
    ])
 
    kpi_row([
        ("NS Current Projection", fmt_pct(ns_o.get("Current Projected %")),
         fmt_cr(ns_o.get("Current FY Estimate")), "off"),
        ("NS Scenario Projection", fmt_pct(ns_o.get("Scenario March %")),
         None if _num(ns_o.get("Scenario Δ pp")) is None
         else fmt_pts(ns_o.get("Scenario Δ pp"))),
        ("NS Current RR", fmt_cr(ns_o.get("Current RR")), "Current pace", "off"),
        ("NS Scenario RR", fmt_cr(ns_o.get("Scenario / Required RR")),
         fmt_pct_signed(ns_o.get("Run Rate Change %"))
         if _num(ns_o.get("Run Rate Change %")) is not None else None),
    ])
 
    st.markdown(
        f"<div class='scenario-highlight'><b>Scenario {model.scenario_id} · "
        f"{model.meta['name']}</b><br>{model.meta['explanation']}</div>",
        unsafe_allow_html=True,
    )
 
    tabs = st.tabs(["Gross Sales · Current vs Scenario", "Net Sales · Current vs Scenario"])
    with tabs[0]:
        _render_metric_bifurcation(gs_frame, gs_formats)
    with tabs[1]:
        _render_metric_bifurcation(ns_frame, ns_formats)
 
    st.markdown(
        "<div class='note'>Overall and Asset Class are shown separately from Channel. "
        "Retail, DHNI and VRM have scenario calculations from their detailed sheets. "
        "Channels such as Insti/Digital remain visible as current FINAL metrics; "
        "scenario fields stay blank where no detailed calculation sheet exists.</div>",
        unsafe_allow_html=True,
    )
 
 
def render_final_dashboard(payload: bytes) -> None:
    """Display the workbook's sixth FINAL sheet inside the Streamlit app."""
    st.markdown(
        "<div class='app-title'>Sales Target Achievement Dashboard</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='app-sub'>Directly read from the workbook's FINAL sheet — "
        "GS / NS achievement, AUM, run-rate estimates and vertical breakouts.</div>",
        unsafe_allow_html=True,
    )
    st.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
 
    try:
        html = build_final_sheet_html(payload)
    except WorkbookError as error:
        st.error(str(error))
        return
 
    st.markdown(
        "<div class='note'>This view is taken from the Excel <b>FINAL</b> sheet. "
        "Use the horizontal scroll inside the dashboard to see all columns.</div>",
        unsafe_allow_html=True,
    )
    st.markdown(html, unsafe_allow_html=True)
 
    with st.expander("View FINAL sheet as raw data", expanded=False):
        try:
            raw = load_final_sheet_frame(payload)
            st.dataframe(raw, use_container_width=True, hide_index=True, height=620)
        except WorkbookError as error:
            st.error(str(error))
 
    st.download_button(
        "Download uploaded workbook",
        data=payload,
        file_name="sales_target_dashboard_source.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
 
 
# =============================================================================
# 4. SEGMENT IDENTIFICATION (Scenario 6)
# =============================================================================
 
def identify_segments(records: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """Suggest a column + values that identify each configured business segment."""
    suggestions: Dict[str, Dict[str, Any]] = {}
    for segment, rule in SEGMENT_RULES.items():
        for column in rule["search_columns"]:
            if column not in records.columns:
                continue
            values = sorted({v for v in text_column(records, column) if v.strip()})
            matches = [
                v for v in values
                if any(keyword in v.casefold() for keyword in rule["keywords"])
            ]
            if matches:
                suggestions[segment] = {"column": column, "values": matches}
                break
    return suggestions
 
 
def map_business_segments(
    records: pd.DataFrame,
    mapping: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    """Assign every record to Digital / Retail B30 / Others using the mapping."""
    out = records.copy()
    out["Segment"] = FALLBACK_SEGMENT
    # Reverse priority so that the first segment in SEGMENT_ORDER wins.
    for segment in reversed([s for s in SEGMENT_ORDER if s != FALLBACK_SEGMENT]):
        rule = mapping.get(segment)
        if not rule:
            continue
        column, values = rule.get("column"), set(rule.get("values") or [])
        if not column or column not in out.columns or not values:
            continue
        mask = text_column(out, column).isin(values)
        out.loc[mask, "Segment"] = segment
    return out
 
 
def segment_diagnostics(records: pd.DataFrame) -> Dict[str, int]:
    counts = records["Segment"].value_counts().to_dict()
    return {segment: int(counts.get(segment, 0)) for segment in SEGMENT_ORDER}
 
 
CHANNEL_KEYWORDS: Dict[str, List[str]] = {
    "Digital": ["digital", "online", "d2c", "e-com", "ecom", "virtual", "web"],
    "VRM": ["vrm", "virtual relationship", "virtual rm"],
    "EM": ["em", "emerging market", "em city"],
    "B30": ["b30", "b-30", "b 30"],
    "T30": ["t30", "t-30", "t 30"],
    "T8": ["t8", "t-8", "t 8"],
    "DHNI": ["dhni", "d-hni", "hni", "wealth"],
    "Retail": ["retail", "rm retail"],
    "Institutional": ["insti", "institutional", "institution", "institutional sales"],
}
 
def _channel_text_score(row: pd.Series, channel: str) -> int:
    values = " ".join(str(row.get(c, "")) for c in META_FIELDS).casefold()
    return sum(1 for keyword in CHANNEL_KEYWORDS[channel] if keyword in values)
 
def identify_channels(records: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """Suggest channel mappings from workbook metadata without hard-coding one schema."""
    suggestions: Dict[str, Dict[str, Any]] = {}
    usable = [f for f in META_FIELDS if f in records.columns and text_column(records, f).ne("").any()]
    for channel in CHANNELS:
        best = (0, None, [])
        for column in usable:
            values = sorted({v for v in text_column(records, column) if v.strip()})
            matches = [v for v in values if any(k in v.casefold() for k in CHANNEL_KEYWORDS[channel])]
            score = len(matches)
            if score > best[0]:
                best = (score, column, matches)
        if best[1] and best[2]:
            suggestions[channel] = {"column": best[1], "values": best[2]}
        elif channel == "VRM" and "Vertical" in records.columns:
            suggestions[channel] = {"column": "Vertical", "values": ["VRM"]}
        elif channel == "DHNI" and "Vertical" in records.columns:
            suggestions[channel] = {"column": "Vertical", "values": ["DHNI"]}
        elif channel == "Retail" and "Vertical" in records.columns:
            suggestions[channel] = {"column": "Vertical", "values": ["Retail"]}
    return suggestions
 
 
def map_business_channels(records: pd.DataFrame, mapping: Dict[str, Dict[str, Any]]) -> pd.DataFrame:
    """Assign every record to one of the nine Scenario 8 planning channels."""
    out = records.copy()
    out["Channel"] = "Unclassified"
    # More specific channels are applied first; explicit mapping wins.
    for channel in CHANNELS:
        rule = mapping.get(channel)
        if not rule:
            continue
        column, values = rule.get("column"), set(rule.get("values") or [])
        if not column or column not in out.columns or not values:
            continue
        mask = text_column(out, column).isin(values)
        out.loc[mask, "Channel"] = channel
    # Use the existing vertical as a safe fallback for the three explicit RM populations.
    if "Vertical" in out.columns:
        for channel, vertical in (("VRM", "VRM"), ("DHNI", "DHNI"), ("Retail", "Retail")):
            mask = (out["Channel"] == "Unclassified") & (out["Vertical"] == vertical)
            out.loc[mask, "Channel"] = channel
    return out
 
 
# =============================================================================
# 5. BASE GRID & CURRENT-STATE STATISTICS
# =============================================================================
 
def build_base_grid(records: pd.DataFrame) -> pd.DataFrame:
    """Aggregate records to the finest analytical grain used by the engine."""
    rows: List[Dict[str, Any]] = []
    grouped = records.groupby(["Vertical", "Segment", "Channel"], dropna=False)
    for (vertical, segment, channel), block in grouped:
        for sales in SALES_TYPES:
            for asset in ASSETS:
                rows.append({
                    "Vertical": vertical,
                    "Segment": segment,
                    "Channel": channel,
                    "Sales": sales,
                    "Asset": asset,
                    "fy_target": float(block[f"{sales}_{asset}_fy"].sum()),
                    "ytd_target": float(block[f"{sales}_{asset}_ytd_tgt"].sum()),
                    "ytd_ach": float(block[f"{sales}_{asset}_ach"].sum()),
                })
    return pd.DataFrame(rows)
 
 
def filter_grid(
    grid: pd.DataFrame,
    sales: Optional[str] = None,
    asset: Optional[str] = None,
    vertical: Optional[str] = None,
    segment: Optional[str] = None,
    channel: Optional[str] = None,
) -> pd.DataFrame:
    mask = pd.Series(True, index=grid.index)
    if sales is not None:
        mask &= grid["Sales"] == sales
    if asset is not None:
        mask &= grid["Asset"] == asset
    if vertical is not None:
        mask &= grid["Vertical"] == vertical
    if segment is not None:
        mask &= grid["Segment"] == segment
    if channel is not None and "Channel" in grid.columns:
        mask &= grid["Channel"] == channel
    return grid.loc[mask]
 
 
def current_asset_stats(fy_target: float, ytd_target: float, ytd_ach: float) -> Dict[str, Any]:
    """Baseline statistics for one asset / group. Never scenario dependent."""
    fy_target = _z(fy_target)
    ytd_target = _z(ytd_target)
    ytd_ach = _z(ytd_ach)
    current_rr = ytd_ach / MONTHS_COMPLETED if MONTHS_COMPLETED else None
    current_march = ytd_ach + (current_rr or 0.0) * MONTHS_REMAINING
    return {
        "fy_target": fy_target,
        "ytd_target": ytd_target,
        "ytd_ach": ytd_ach,
        "current_rr": current_rr,
        "ytd_ach_pct": safe_div(ytd_ach, ytd_target),
        "fy_completed_pct": safe_div(ytd_ach, fy_target),
        "current_march": current_march,
        "current_march_pct": safe_div(current_march, fy_target),
    }
 
 
def summarize_current(grid: pd.DataFrame, **filters: Any) -> Dict[str, Any]:
    """Baseline statistics for an arbitrary slice of the base grid."""
    subset = filter_grid(grid, **filters)
    return current_asset_stats(
        subset["fy_target"].sum(),
        subset["ytd_target"].sum(),
        subset["ytd_ach"].sum(),
    )
 
 
# =============================================================================
# 6. SCENARIO ENGINE - SCENARIOS 1 TO 6
# =============================================================================
 
def _blank_cell(stats: Dict[str, Any]) -> Dict[str, Any]:
    cell = dict(stats)
    cell.update({
        "scen_rr": None, "rr_change_pct": None, "feb_mar_rr": None,
        "jan_required": None, "jan_amount": None, "jan_pct": None,
        "jan_buffer": None, "jan_buffer_pct": None,
        "march_required": None, "march_amount": None, "march_pct": None,
        "milestone_pct": None, "incremental_sales": None,
        "headroom_amt": None, "headroom_pct": None,
        "momentum_g": None, "feasible": None, "binding": None,
        "trajectory": None, "note": "",
    })
    return cell
 
 
def compute_cell(
    fy_target: float,
    ytd_target: float,
    ytd_ach: float,
    kind: str,
    multiplier: Optional[float] = None,
    uplift: Optional[float] = None,
    dip: float = 0.0,
) -> Dict[str, Any]:
    """Scenario mathematics for one asset / group (scenarios 1-6)."""
    stats = current_asset_stats(fy_target, ytd_target, ytd_ach)
    cell = _blank_cell(stats)
    ach = stats["ytd_ach"]
    current_rr = stats["current_rr"] or 0.0
 
    if kind == "runrate":
        scen_rr = current_rr * (1.0 + (uplift or 0.0))
        jan_amount = ach + scen_rr * MONTHS_JUL_JAN
        cell.update({
            "scen_rr": scen_rr,
            "feb_mar_rr": scen_rr,
            "jan_amount": jan_amount,
            "march_amount": ach + scen_rr * MONTHS_REMAINING,
            "milestone_pct": None,
        })
 
    elif kind == "jan_target":
        required = max(_z(multiplier) * stats["fy_target"], ach)
        scen_rr = max(required - ach, 0.0) / MONTHS_JUL_JAN
        jan_amount = ach + scen_rr * MONTHS_JUL_JAN
        feb_mar_rr = scen_rr * (1.0 - dip)
        cell.update({
            "scen_rr": scen_rr,
            "feb_mar_rr": feb_mar_rr,
            "jan_required": required,
            "jan_amount": jan_amount,
            "march_amount": jan_amount + feb_mar_rr * MONTHS_FEB_MAR,
            "milestone_pct": multiplier,
        })
 
    elif kind == "march_target":
        required = max(_z(multiplier) * stats["fy_target"], ach)
        scen_rr = max(required - ach, 0.0) / MONTHS_REMAINING
        cell.update({
            "scen_rr": scen_rr,
            "feb_mar_rr": scen_rr,
            "march_required": required,
            "jan_amount": ach + scen_rr * MONTHS_JUL_JAN,
            "march_amount": ach + scen_rr * MONTHS_REMAINING,
            "milestone_pct": multiplier,
        })
 
    else:  # pragma: no cover - guarded by the scenario registry
        raise ValueError(f"Unknown scenario kind: {kind}")
 
    return _finalise_cell(cell)
 
 
def _finalise_cell(cell: Dict[str, Any]) -> Dict[str, Any]:
    """Recompute all derived ratios from the absolute amounts in the cell."""
    fy_target = cell.get("fy_target")
    current_rr = _num(cell.get("current_rr"))
    scen_rr = _num(cell.get("scen_rr"))
 
    cell["ytd_ach_pct"] = safe_div(cell.get("ytd_ach"), cell.get("ytd_target"))
    cell["fy_completed_pct"] = safe_div(cell.get("ytd_ach"), fy_target)
    cell["current_march_pct"] = safe_div(cell.get("current_march"), fy_target)
    cell["jan_pct"] = safe_div(cell.get("jan_amount"), fy_target)
    cell["march_pct"] = safe_div(cell.get("march_amount"), fy_target)
 
    if current_rr is not None and current_rr > 0 and scen_rr is not None:
        cell["rr_change_pct"] = (scen_rr / current_rr) - 1.0
    else:
        cell["rr_change_pct"] = None
 
    march_amount = _num(cell.get("march_amount"))
    current_march = _num(cell.get("current_march"))
    if march_amount is not None and current_march is not None:
        cell["incremental_sales"] = march_amount - current_march
 
    jan_required = _num(cell.get("jan_required"))
    jan_amount = _num(cell.get("jan_amount"))
    if jan_required is not None and jan_amount is not None:
        cell["jan_buffer"] = jan_amount - jan_required
        cell["jan_buffer_pct"] = safe_div(cell["jan_buffer"], jan_required)
 
    march_required = _num(cell.get("march_required"))
    if march_required is not None and march_amount is not None:
        cell["headroom_amt"] = march_amount - march_required
        march_pct = cell.get("march_pct")
        required_pct = safe_div(march_required, fy_target)
        if march_pct is not None and required_pct is not None:
            cell["headroom_pct"] = march_pct - required_pct
        cell["feasible"] = cell["headroom_amt"] >= -1e-6
 
    if cell.get("milestone_pct") is None:
        cell["milestone_pct"] = cell.get("march_pct")
    return cell
 
 
def scenario_multipliers(grid: pd.DataFrame, scenario_id: int) -> Dict[Tuple[str, str, str], float]:
    """
    Derive the per-asset FY-target multiplier for the selected scenario.
 
    Key is (sales type, asset, segment); segment is '*' unless the scenario
    differentiates by business segment.
    """
    multipliers: Dict[Tuple[str, str, str], float] = {}
 
    if scenario_id in (1, 7):
        return multipliers
 
    if scenario_id == 3:
        for sales in SALES_TYPES:
            for asset in ASSETS:
                multipliers[(sales, asset, "*")] = S3_TARGET
        return multipliers
 
    if scenario_id == 4:
        for sales in SALES_TYPES:
            for asset in ASSETS:
                multipliers[(sales, asset, "*")] = S4_TARGET
        return multipliers
 
    if scenario_id == 6:
        for sales in SALES_TYPES:
            for asset in ASSETS:
                for segment in SEGMENT_ORDER:
                    multipliers[(sales, asset, segment)] = S6_SEGMENT_TARGETS.get(segment, 1.0)
        return multipliers
 
    # Scenarios 2 and 5 balance Debt and Liquid around a fixed Equity ambition.
    equity_mult = S2_EQUITY_TARGET if scenario_id == 2 else S5_EQUITY_TARGET
    overall_mult = S2_OVERALL_TARGET if scenario_id == 2 else S5_OVERALL_TARGET
 
    for sales in SALES_TYPES:
        targets = {
            asset: float(filter_grid(grid, sales=sales, asset=asset)["fy_target"].sum())
            for asset in ASSETS
        }
        total_target = sum(targets.values())
        required_overall = overall_mult * total_target
        required_equity = equity_mult * targets["Equity"]
        remaining = max(required_overall - required_equity, 0.0)
        denominator = targets["Debt"] + targets["Liquid"]
        share = safe_div(remaining, denominator)
        balance_mult = 0.0 if share is None else share
        multipliers[(sales, "Equity", "*")] = equity_mult
        multipliers[(sales, "Debt", "*")] = balance_mult
        multipliers[(sales, "Liquid", "*")] = balance_mult
 
    return multipliers
 
 
def _multiplier_for(
    multipliers: Dict[Tuple[str, str, str], float],
    sales: str,
    asset: str,
    segment: str,
) -> Optional[float]:
    if (sales, asset, segment) in multipliers:
        return multipliers[(sales, asset, segment)]
    return multipliers.get((sales, asset, "*"))
 
 
def apply_scenario_grid(
    grid: pd.DataFrame,
    scenario_id: int,
    params: Dict[str, Any],
    multipliers: Dict[Tuple[str, str, str], float],
) -> pd.DataFrame:
    """Evaluate scenarios 1-6 over every cell of the base grid."""
    kind = SCENARIOS[scenario_id]["kind"]
    dip = float(params.get("dip", 0.0)) if scenario_id == 3 else 0.0
    uplift = S1_RUNRATE_UPLIFT if scenario_id == 1 else None
 
    results: List[Dict[str, Any]] = []
    for row in grid.to_dict("records"):
        multiplier = _multiplier_for(multipliers, row["Sales"], row["Asset"], row["Segment"])
        cell = compute_cell(
            row["fy_target"], row["ytd_target"], row["ytd_ach"],
            kind=kind, multiplier=multiplier, uplift=uplift, dip=dip,
        )
        cell.update({
            "Vertical": row["Vertical"], "Segment": row["Segment"],
            "Sales": row["Sales"], "Asset": row["Asset"],
        })
        results.append(cell)
    return pd.DataFrame(results)
 
 
SUMMABLE_FIELDS = [
    "fy_target", "ytd_target", "ytd_ach", "current_rr", "current_march",
    "scen_rr", "feb_mar_rr", "jan_required", "jan_amount",
    "march_required", "march_amount",
]
 
 
def summarize_cells(subset: pd.DataFrame) -> Dict[str, Any]:
    """Aggregate scenario cells additively and rebuild every derived ratio."""
    cell: Dict[str, Any] = {}
    for field in SUMMABLE_FIELDS:
        cell[field] = _ssum(subset[field]) if field in subset.columns else None
    milestones = subset["milestone_pct"].dropna().unique() if "milestone_pct" in subset else []
    cell["milestone_pct"] = float(milestones[0]) if len(milestones) == 1 else None
    cell["note"] = ""
    cell["momentum_g"] = None
    cell["trajectory"] = None
    cell["binding"] = None
    cell["feasible"] = None
    return _finalise_cell(cell)
 
 
# --- Named scenario entry points (thin wrappers over the shared engine) -------
 
def _scenario_frame(grid: pd.DataFrame, scenario_id: int, params: Dict[str, Any]) -> pd.DataFrame:
    return apply_scenario_grid(grid, scenario_id, params, scenario_multipliers(grid, scenario_id))
 
 
def calculate_scenario_1(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """+20% run-rate push from July onward."""
    return _scenario_frame(grid, 1, params)
 
 
def calculate_scenario_2(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """100% Equity and 75% overall FY target by January."""
    return _scenario_frame(grid, 2, params)
 
 
def calculate_scenario_3(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """100% of FY target by January, then a configurable Feb-Mar dip."""
    return _scenario_frame(grid, 3, params)
 
 
def calculate_scenario_4(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """120% of FY target by March."""
    return _scenario_frame(grid, 4, params)
 
 
def calculate_scenario_5(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """120% Equity and 100% overall FY target by March."""
    return _scenario_frame(grid, 5, params)
 
 
def calculate_scenario_6(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """Digital 140%, Retail B30 125%, Others 115% of their FY targets."""
    return _scenario_frame(grid, 6, params)
 
 
SCENARIO_FUNCTIONS = {
    1: calculate_scenario_1,
    2: calculate_scenario_2,
    3: calculate_scenario_3,
    4: calculate_scenario_4,
    5: calculate_scenario_5,
    6: calculate_scenario_6,
}
 
 
# =============================================================================
# 7. SCENARIO 7 - MOMENTUM ENGINE
# =============================================================================
 
def _momentum_sum(growth: float, months: int, tail_factors: Sequence[float]) -> float:
    """Sum of compounding monthly run-rate multiples, including leakage tail."""
    factor = 1.0 + growth
    build = sum(factor ** k for k in range(1, months + 1))
    tail = sum(f * factor ** months for f in tail_factors)
    return build + tail
 
 
def solve_momentum_rate(
    current_rr: float,
    required_amount: float,
    months: int = MONTHS_JUL_JAN,
    tail_factors: Sequence[float] = (),
    upper: float = 3.0,
) -> Optional[float]:
    """
    Back-solve the minimum month-on-month growth rate g such that the
    compounding trajectory delivers the required incremental amount.
 
    Returns 0.0 when no additional momentum is needed and None when the
    requirement cannot be met within the search bounds.
    """
    rr = _num(current_rr)
    need = _num(required_amount)
    if rr is None or need is None or rr <= 0:
        return None
    if need <= 0:
        return 0.0
 
    def shortfall(growth: float) -> float:
        return rr * _momentum_sum(growth, months, tail_factors) - need
 
    if shortfall(0.0) >= 0:
        return 0.0
    if shortfall(upper) < 0:
        return None
 
    low, high = 0.0, upper
    for _ in range(240):
        mid = (low + high) / 2.0
        if shortfall(mid) >= 0:
            high = mid
        else:
            low = mid
    return high
 
 
def calculate_momentum_trajectory(
    current_rr: float,
    growth: Optional[float],
    leakage: float,
    flat_rate: Optional[float] = None,
) -> List[float]:
    """Monthly run rates for July → March (momentum build, then leakage)."""
    if growth is None:
        base = _z(flat_rate)
        build = [base] * MONTHS_JUL_JAN
    else:
        rr = _z(current_rr)
        build = [rr * (1.0 + growth) ** k for k in range(1, MONTHS_JUL_JAN + 1)]
    january_rr = build[-1] if build else 0.0
    february_rr = january_rr * (1.0 - leakage)
    march_rr = february_rr * (1.0 - leakage)
    return build + [february_rr, march_rr]
 
 
def calculate_leakage_impact(january_rr: float, leakage: float) -> Dict[str, float]:
    """February and March run rates after AUM leakage / run-rate pressure."""
    february_rr = _z(january_rr) * (1.0 - leakage)
    march_rr = february_rr * (1.0 - leakage)
    return {
        "february_rr": february_rr,
        "march_rr": march_rr,
        "feb_mar_sales": february_rr + march_rr,
    }
 
 
def calculate_momentum_headroom(
    march_amount: float,
    march_required: float,
    fy_target: float,
    march_target_pct: float,
) -> Dict[str, Optional[float]]:
    """Scenario achievement versus the March ambition, in ₹ and in points."""
    headroom_amt = _z(march_amount) - _z(march_required)
    achieved_pct = safe_div(march_amount, fy_target)
    headroom_pct = None if achieved_pct is None else achieved_pct - march_target_pct
    return {"headroom_amt": headroom_amt, "headroom_pct": headroom_pct}
 
 
def calculate_scenario_7(
    fy_target: float,
    ytd_target: float,
    ytd_ach: float,
    params: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Momentum build-up model.
 
    Solves for the month-on-month growth rate that simultaneously satisfies the
    January milestone and, after Feb-Mar leakage, the March ambition.
    """
    jan_target_pct = float(params.get("jan_target", S7_DEFAULT_JAN_TARGET))
    mar_target_pct = float(params.get("mar_target", S7_DEFAULT_MAR_TARGET))
    leakage = float(params.get("leakage", S7_DEFAULT_LEAKAGE))
 
    stats = current_asset_stats(fy_target, ytd_target, ytd_ach)
    cell = _blank_cell(stats)
    ach = stats["ytd_ach"]
    current_rr = stats["current_rr"] or 0.0
 
    jan_required = max(jan_target_pct * stats["fy_target"], ach)
    mar_required = max(mar_target_pct * stats["fy_target"], ach)
    tail = ((1.0 - leakage), (1.0 - leakage) ** 2)
 
    growth_jan = solve_momentum_rate(current_rr, jan_required - ach, MONTHS_JUL_JAN)
    growth_mar = solve_momentum_rate(current_rr, mar_required - ach, MONTHS_JUL_JAN, tail)
 
    note = ""
    binding = None
    flat_rate = None
 
    if current_rr <= 0:
        # Momentum compounding is undefined on a non-positive run rate:
        # fall back to the flat run rate required to hold both milestones.
        flat_jan = max(jan_required - ach, 0.0) / MONTHS_JUL_JAN
        denominator = MONTHS_JUL_JAN + tail[0] + tail[1]
        flat_mar = max(mar_required - ach, 0.0) / denominator
        flat_rate = max(flat_jan, flat_mar)
        growth = None
        binding = "January" if flat_jan >= flat_mar else "March"
        note = (
            "Current run rate is not positive, so compounding momentum cannot be applied. "
            "The flat monthly run rate required to hold the milestones is shown instead."
        )
    elif growth_jan is None and growth_mar is None:
        growth = None
        flat_rate = max(mar_required - ach, 0.0) / (MONTHS_JUL_JAN + tail[0] + tail[1])
        binding = "March"
        note = "The requirement exceeds the momentum search range; a flat required run rate is shown."
    else:
        candidates = [g for g in (growth_jan, growth_mar) if g is not None]
        growth = max(candidates)
        binding = "March" if (growth_mar is not None and growth == growth_mar
                              and (growth_jan is None or growth_mar >= growth_jan)) else "January"
 
    trajectory = calculate_momentum_trajectory(current_rr, growth, leakage, flat_rate)
    build_phase = trajectory[:MONTHS_JUL_JAN]
    jan_amount = ach + sum(build_phase)
    january_rr = build_phase[-1] if build_phase else 0.0
    leak = calculate_leakage_impact(january_rr, leakage)
    march_amount = jan_amount + leak["feb_mar_sales"]
 
    headroom = calculate_momentum_headroom(
        march_amount, mar_required, stats["fy_target"], mar_target_pct
    )
    shortfall = max(mar_required - march_amount, 0.0)
    denominator = 1.0 + tail[0] + tail[1]
    additional_jan_rr = shortfall / denominator if denominator else None
 
    cell.update({
        "scen_rr": january_rr,
        "feb_mar_rr": leak["february_rr"],
        "march_rr": leak["march_rr"],
        "jan_required": jan_required,
        "jan_amount": jan_amount,
        "march_required": mar_required,
        "march_amount": march_amount,
        "milestone_pct": jan_target_pct,
        "march_target_pct": mar_target_pct,
        "momentum_g": growth,
        "flat_rate": flat_rate,
        "leakage": leakage,
        "trajectory": trajectory,
        "binding": binding,
        "note": note,
        "additional_march_sales": shortfall,
        "additional_jan_rr": additional_jan_rr,
        "avg_scen_rr": (sum(trajectory) / len(trajectory)) if trajectory else None,
    })
    cell = _finalise_cell(cell)
    cell["headroom_amt"] = headroom["headroom_amt"]
    cell["headroom_pct"] = headroom["headroom_pct"]
    cell["feasible"] = shortfall <= 1e-6
    # Momentum run rate versus the current flat run rate, measured on the
    # January exit rate (the pace the business must be running at by then).
    cell["rr_change_pct"] = (
        (january_rr / current_rr) - 1.0 if current_rr and current_rr > 0 else None
    )
    return cell
 
 
# =============================================================================
# 7A. SCENARIO 8/9 - CHANNEL SIMULATOR & MIX OPTIMISER
# =============================================================================
 
def _s8_channel_params(params: Dict[str, Any], channel: str) -> Tuple[float, float, float, float]:
    growth = float(params.get("channel_growth", {}).get(channel, S8_DEFAULT_GROWTH[channel]))
    jan_target = float(params.get("channel_jan_target", {}).get(channel, S8_DEFAULT_JAN_TARGET[channel]))
    mar_target = float(params.get("channel_mar_target", {}).get(channel, S8_DEFAULT_MAR_TARGET[channel]))
    leakage = float(params.get("leakage", S8_DEFAULT_LEAKAGE))
    return growth, jan_target, mar_target, leakage
 
 
def calculate_scenario_8_cell(fy_target: float, ytd_target: float, ytd_ach: float, channel: str, params: Dict[str, Any]) -> Dict[str, Any]:
    """Fixed-growth channel trajectory with independent Jan-2027 and Mar-2027 targets."""
    stats = current_asset_stats(fy_target, ytd_target, ytd_ach)
    cell = _blank_cell(stats)
    growth, jan_target_pct, mar_target_pct, leakage = _s8_channel_params(params, channel)
    current_rr = _z(stats["current_rr"])
    trajectory = calculate_momentum_trajectory(current_rr, growth, leakage)
    jan_amount = _z(stats["ytd_ach"]) + sum(trajectory[:MONTHS_JUL_JAN])
    march_amount = jan_amount + sum(trajectory[MONTHS_JUL_JAN:])
    jan_required = max(jan_target_pct * stats["fy_target"], _z(stats["ytd_ach"]))
    march_required = max(mar_target_pct * stats["fy_target"], _z(stats["ytd_ach"]))
    jan_gap = jan_amount - jan_required
    march_gap = march_amount - march_required
    cell.update({
        "channel": channel, "scen_rr": trajectory[MONTHS_JUL_JAN-1] if trajectory else current_rr,
        "feb_mar_rr": trajectory[MONTHS_JUL_JAN] if len(trajectory) > MONTHS_JUL_JAN else None,
        "march_rr": trajectory[-1] if trajectory else None,
        "jan_required": jan_required, "jan_amount": jan_amount,
        "march_required": march_required, "march_amount": march_amount,
        "milestone_pct": jan_target_pct, "march_target_pct": mar_target_pct,
        "momentum_g": growth, "leakage": leakage, "trajectory": trajectory,
        "jan_buffer": jan_gap, "jan_buffer_pct": safe_div(jan_gap, jan_required),
        "headroom_amt": march_gap, "headroom_pct": safe_div(march_gap, march_required),
        "feasible": jan_gap >= -1e-6 and march_gap >= -1e-6,
        "binding": "January" if jan_gap < 0 else ("March" if march_gap < 0 else "None"),
        "additional_march_sales": max(-march_gap, 0.0),
        "additional_jan_rr": max(-jan_gap, 0.0) / MONTHS_JUL_JAN,
    })
    return _finalise_cell(cell)
 
 
def calculate_scenario_8_grid(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for row in grid.to_dict("records"):
        channel = row.get("Channel", "Unclassified")
        cell = calculate_scenario_8_cell(row["fy_target"], row["ytd_target"], row["ytd_ach"], channel, params)
        cell.update({k: row[k] for k in ("Vertical", "Segment", "Channel", "Sales", "Asset")})
        rows.append(cell)
    return pd.DataFrame(rows)
 
 
def calculate_scenario_9_grid(grid: pd.DataFrame, params: Dict[str, Any]) -> pd.DataFrame:
    """Optimiser: solve minimum MoM growth per channel to meet its Jan/Mar targets."""
    rows = []
    for row in grid.to_dict("records"):
        channel = row.get("Channel", "Unclassified")
        _, jan_target, mar_target, leakage = _s8_channel_params(params, channel)
        stats = current_asset_stats(row["fy_target"], row["ytd_target"], row["ytd_ach"])
        ach = _z(stats["ytd_ach"]); rr = _z(stats["current_rr"])
        jan_req = max(jan_target * stats["fy_target"], ach)
        mar_req = max(mar_target * stats["fy_target"], ach)
        gj = solve_momentum_rate(rr, jan_req-ach, MONTHS_JUL_JAN)
        gm = solve_momentum_rate(rr, mar_req-ach, MONTHS_JUL_JAN, ((1-leakage), (1-leakage)**2))
        growth = max([g for g in (gj, gm) if g is not None], default=0.0)
        cell = calculate_scenario_8_cell(row["fy_target"], row["ytd_target"], row["ytd_ach"], channel, {
            **params, "channel_growth": {**params.get("channel_growth", {}), channel: growth}
        })
        cell.update({k: row[k] for k in ("Vertical", "Segment", "Channel", "Sales", "Asset")})
        cell["optimized_growth"] = growth
        rows.append(cell)
    return pd.DataFrame(rows)
 
 
def build_channel_scenario_analysis(model: "ScenarioModel", basis: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    rows = []
    frame = model.scenario_grid
    for channel in CHANNELS:
        subset = frame[(frame["Sales"] == basis) & (frame["Channel"] == channel)]
        if subset.empty:
            continue
        cell = summarize_cells(subset)
        growth, jan_target, mar_target, leakage = _s8_channel_params(model.params, channel)
        rows.append({
            "Channel": channel, "MoM Growth": growth,
            "Jan 2027 Target": jan_target, "Jan Achievement": cell.get("jan_pct"),
            "Jan Gap / Headroom": cell.get("jan_buffer"),
            "Mar 2027 Target": mar_target, "Mar Achievement": cell.get("march_pct"),
            "Mar Gap / Headroom": cell.get("headroom_amt"),
            "Current Run Rate": cell.get("current_rr"),
            "Jan Exit Run Rate": cell.get("scen_rr"),
            "March Incremental Sales": cell.get("incremental_sales"),
        })
    formats = {
        "Channel":"txt", "MoM Growth":"pct_signed", "Jan 2027 Target":"pct",
        "Jan Achievement":"pct", "Jan Gap / Headroom":"cr_signed",
        "Mar 2027 Target":"pct", "Mar Achievement":"pct", "Mar Gap / Headroom":"cr_signed",
        "Current Run Rate":"cr", "Jan Exit Run Rate":"cr", "March Incremental Sales":"cr_signed"
    }
    return pd.DataFrame(rows), formats
 
 
# =============================================================================
# 8. SCENARIO MODEL - ONE INTERFACE FOR EVERY VIEW
# =============================================================================
 
class ScenarioModel:
    """Evaluates the selected scenario for any slice of the business."""
 
    def __init__(self, scenario_id: int, grid: pd.DataFrame, params: Dict[str, Any]):
        self.scenario_id = scenario_id
        self.meta = SCENARIOS[scenario_id]
        self.grid = grid
        self.params = params
        self.multipliers = scenario_multipliers(grid, scenario_id)
        self._cache: Dict[Tuple, Dict[str, Any]] = {}
        if scenario_id == 7:
            self.scenario_grid = None
        elif scenario_id == 8:
            self.scenario_grid = calculate_scenario_8_grid(grid, params)
        elif scenario_id == 9:
            self.scenario_grid = calculate_scenario_9_grid(grid, params)
        else:
            self.scenario_grid = SCENARIO_FUNCTIONS[scenario_id](grid, params)
 
    # -- core accessor --------------------------------------------------------
    def cell(
        self,
        sales: str,
        asset: Optional[str] = None,
        vertical: Optional[str] = None,
        segment: Optional[str] = None,
        channel: Optional[str] = None,
    ) -> Dict[str, Any]:
        key = (sales, asset, vertical, segment)
        if key in self._cache:
            return self._cache[key]
 
        if self.scenario_id == 7:
            subset = filter_grid(self.grid, sales=sales, asset=asset,
                                 vertical=vertical, segment=segment, channel=channel)
            cell = calculate_scenario_7(
                subset["fy_target"].sum(),
                subset["ytd_target"].sum(),
                subset["ytd_ach"].sum(),
                self.params,
            )
        else:
            frame = self.scenario_grid
            mask = frame["Sales"] == sales
            if asset is not None:
                mask &= frame["Asset"] == asset
            if vertical is not None:
                mask &= frame["Vertical"] == vertical
            if segment is not None:
                mask &= frame["Segment"] == segment
            if channel is not None and "Channel" in frame.columns:
                mask &= frame["Channel"] == channel
            cell = summarize_cells(frame.loc[mask])
 
        self._cache[key] = cell
        return cell
 
    # -- convenience views ----------------------------------------------------
    def assets(self, sales: str, **filters: Any) -> Dict[str, Dict[str, Any]]:
        return {asset: self.cell(sales, asset=asset, **filters) for asset in ASSETS}
 
    def baseline(self, sales: str, **filters: Any) -> Dict[str, Any]:
        return summarize_current(self.grid, sales=sales, **filters)
 
    def implied_milestones(self, sales: str) -> Dict[str, Optional[float]]:
        return {
            asset: _multiplier_for(self.multipliers, sales, asset, "*")
            for asset in ASSETS
        }
 
    def available_segments(self) -> List[str]:
        present = set(self.grid["Segment"].unique())
        return [s for s in SEGMENT_ORDER if s in present]
 
    def available_verticals(self) -> List[str]:
        present = set(self.grid["Vertical"].unique())
        return [v for v in VERTICALS if v in present]
 
 
# =============================================================================
# 9. REVENUE ENGINE
# =============================================================================
 
def calculate_revenue(cells_by_asset: Dict[str, Dict[str, Any]], field: str) -> Dict[str, Any]:
    """Asset-class revenue from a set of scenario cells. No blended rate."""
    by_asset: Dict[str, Optional[float]] = {}
    sales_by_asset: Dict[str, Optional[float]] = {}
    total = 0.0
    for asset in ASSETS:
        amount = _num(cells_by_asset.get(asset, {}).get(field))
        sales_by_asset[asset] = amount
        revenue = None if amount is None else amount * REVENUE_RATE[asset]
        by_asset[asset] = revenue
        total += 0.0 if revenue is None else revenue
    return {"by_asset": by_asset, "sales_by_asset": sales_by_asset, "total": total}
 
 
def calculate_baseline_revenue(cells_by_asset: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """Baseline revenue on the current-run-rate March projection (scenario independent)."""
    return calculate_revenue(cells_by_asset, "current_march")
 
 
def calculate_incremental_revenue(
    scenario_revenue: Dict[str, Any],
    baseline_revenue: Dict[str, Any],
) -> Dict[str, Any]:
    incremental = {
        asset: (
            None
            if scenario_revenue["by_asset"].get(asset) is None
            or baseline_revenue["by_asset"].get(asset) is None
            else scenario_revenue["by_asset"][asset] - baseline_revenue["by_asset"][asset]
        )
        for asset in ASSETS
    }
    total = scenario_revenue["total"] - baseline_revenue["total"]
    uplift = safe_div(scenario_revenue["total"], baseline_revenue["total"])
    return {
        "by_asset": incremental,
        "total": total,
        "uplift_pct": None if uplift is None else uplift - 1.0,
        "contribution": {
            asset: safe_div(scenario_revenue["by_asset"].get(asset), scenario_revenue["total"])
            for asset in ASSETS
        },
    }
 
 
def revenue_bundle(model: ScenarioModel, basis: str, **filters: Any) -> Dict[str, Any]:
    """Baseline / scenario / incremental revenue for any slice of the business."""
    cells = model.assets(basis, **filters)
    baseline = calculate_baseline_revenue(cells)
    scenario = calculate_revenue(cells, "march_amount")
    january = calculate_revenue(cells, "jan_amount")
    incremental = calculate_incremental_revenue(scenario, baseline)
    return {
        "cells": cells,
        "baseline": baseline,
        "scenario": scenario,
        "january": january,
        "incremental": incremental,
    }
 
 
# =============================================================================
# 10. TABLE BUILDERS
# =============================================================================
 
def build_current_overview(model: ScenarioModel) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Baseline table - identical for every scenario."""
    rows = []
    for sales in SALES_TYPES:
        base = model.baseline(sales)
        rows.append({
            "Sales": SALES_LABEL[sales],
            "FY Target": base["fy_target"],
            "YTD June Target": base["ytd_target"],
            "YTD June Achievement": base["ytd_ach"],
            "Target Achieved %": base["ytd_ach_pct"],
            "FY Target Completed %": base["fy_completed_pct"],
            "Current Run Rate": base["current_rr"],
            "Current March Projection": base["current_march"],
            "Current March Projection %": base["current_march_pct"],
        })
    formats = {
        "Sales": "txt", "FY Target": "cr", "YTD June Target": "cr",
        "YTD June Achievement": "cr", "Target Achieved %": "pct",
        "FY Target Completed %": "pct", "Current Run Rate": "cr",
        "Current March Projection": "cr", "Current March Projection %": "pct",
    }
    return pd.DataFrame(rows), formats
 
 
def summarize_scenario(model: ScenarioModel, sales: str) -> Dict[str, Any]:
    """Headline current-versus-scenario numbers for one sales basis."""
    cell = model.cell(sales)
    return {
        "Sales": SALES_LABEL[sales],
        "Current Run Rate": cell["current_rr"],
        "Scenario Run Rate": cell["scen_rr"],
        "Run Rate Change %": cell["rr_change_pct"],
        "Jan Achievement": cell["jan_amount"],
        "Jan Achievement %": cell["jan_pct"],
        "Feb-Mar Run Rate": cell["feb_mar_rr"],
        "Current March Projection %": cell["current_march_pct"],
        "Scenario March Achievement": cell["march_amount"],
        "Scenario March Achievement %": cell["march_pct"],
        "Incremental Sales": cell["incremental_sales"],
    }
 
 
def build_comparison(model: ScenarioModel) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Current versus selected scenario, for Gross Sales and Net Sales."""
    rows = [summarize_scenario(model, sales) for sales in SALES_TYPES]
    formats = {
        "Sales": "txt", "Current Run Rate": "cr", "Scenario Run Rate": "cr",
        "Run Rate Change %": "pct_signed", "Jan Achievement": "cr",
        "Jan Achievement %": "pct", "Feb-Mar Run Rate": "cr",
        "Current March Projection %": "pct", "Scenario March Achievement": "cr",
        "Scenario March Achievement %": "pct", "Incremental Sales": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_revenue_impact(model: ScenarioModel, basis: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Revenue by asset class - never a blended rate."""
    bundle = revenue_bundle(model, basis)
    rows = []
    for asset in ASSETS:
        rows.append({
            "Asset Class": asset,
            "Scenario Sales": bundle["scenario"]["sales_by_asset"][asset],
            "Revenue Rate": f"{REVENUE_BPS[asset]:.0f} bps",
            "Scenario Revenue": bundle["scenario"]["by_asset"][asset],
            "Baseline Revenue": bundle["baseline"]["by_asset"][asset],
            "Incremental Revenue": bundle["incremental"]["by_asset"][asset],
            "Revenue Contribution %": bundle["incremental"]["contribution"][asset],
        })
    rows.append({
        "Asset Class": "Total",
        "Scenario Sales": sum(_z(v) for v in bundle["scenario"]["sales_by_asset"].values()),
        "Revenue Rate": "—",
        "Scenario Revenue": bundle["scenario"]["total"],
        "Baseline Revenue": bundle["baseline"]["total"],
        "Incremental Revenue": bundle["incremental"]["total"],
        "Revenue Contribution %": 1.0 if bundle["scenario"]["total"] else None,
    })
    formats = {
        "Asset Class": "txt", "Scenario Sales": "cr", "Revenue Rate": "txt",
        "Scenario Revenue": "cr1", "Baseline Revenue": "cr1",
        "Incremental Revenue": "cr1_signed", "Revenue Contribution %": "pct",
    }
    return pd.DataFrame(rows), formats
 
 
def build_vertical_summary(model: ScenarioModel) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Retail, DHNI and VRM, for Gross Sales and Net Sales."""
    rows = []
    for vertical in model.available_verticals():
        for sales in SALES_TYPES:
            cell = model.cell(sales, vertical=vertical)
            bundle = revenue_bundle(model, sales, vertical=vertical)
            rows.append({
                "Vertical": vertical,
                "Sales": SALES_LABEL[sales],
                "FY Target": cell["fy_target"],
                "YTD Achievement": cell["ytd_ach"],
                "Target Achieved %": cell["ytd_ach_pct"],
                "Current Run Rate": cell["current_rr"],
                "Scenario Run Rate": cell["scen_rr"],
                "Run Rate Change %": cell["rr_change_pct"],
                "Current March Projection %": cell["current_march_pct"],
                "Scenario Milestone %": cell["milestone_pct"],
                "Scenario March Projection %": cell["march_pct"],
                "Scenario Revenue": bundle["scenario"]["total"],
                "Incremental Revenue": bundle["incremental"]["total"],
            })
    formats = {
        "Vertical": "txt", "Sales": "txt", "FY Target": "cr", "YTD Achievement": "cr",
        "Target Achieved %": "pct", "Current Run Rate": "cr", "Scenario Run Rate": "cr",
        "Run Rate Change %": "pct_signed", "Current March Projection %": "pct",
        "Scenario Milestone %": "pct", "Scenario March Projection %": "pct",
        "Scenario Revenue": "cr1", "Incremental Revenue": "cr1_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_asset_breakdown(model: ScenarioModel, sales: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Equity / Debt / Liquid split within each vertical."""
    rows = []
    include_dip = model.scenario_id in (3, 7)
    for vertical in model.available_verticals():
        for asset in ASSETS:
            cell = model.cell(sales, asset=asset, vertical=vertical)
            revenue = _z(cell["march_amount"]) * REVENUE_RATE[asset]
            baseline_revenue = _z(cell["current_march"]) * REVENUE_RATE[asset]
            row = {
                "Vertical": vertical,
                "Asset": asset,
                "FY Target": cell["fy_target"],
                "YTD Achievement": cell["ytd_ach"],
                "Target Achieved %": cell["ytd_ach_pct"],
                "Current Run Rate": cell["current_rr"],
                "Scenario Run Rate": cell["scen_rr"],
                "Run Rate Change %": cell["rr_change_pct"],
            }
            if include_dip:
                row["Feb-Mar Run Rate"] = cell["feb_mar_rr"]
            row.update({
                "Current March Projection %": cell["current_march_pct"],
                "Scenario Milestone %": cell["milestone_pct"],
                "Scenario March Projection %": cell["march_pct"],
                "Scenario Revenue": revenue,
                "Incremental Revenue": revenue - baseline_revenue,
            })
            rows.append(row)
    formats = {
        "Vertical": "txt", "Asset": "txt", "FY Target": "cr", "YTD Achievement": "cr",
        "Target Achieved %": "pct", "Current Run Rate": "cr", "Scenario Run Rate": "cr",
        "Run Rate Change %": "pct_signed", "Feb-Mar Run Rate": "cr",
        "Current March Projection %": "pct", "Scenario Milestone %": "pct",
        "Scenario March Projection %": "pct", "Scenario Revenue": "cr1",
        "Incremental Revenue": "cr1_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_segment_scenario_analysis(
    model: ScenarioModel, sales: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 6 - differentiated performance by business segment."""
    rows = []
    for segment in model.available_segments():
        cell = model.cell(sales, segment=segment)
        rows.append({
            "Segment": segment,
            "FY Target": cell["fy_target"],
            "YTD Achievement": cell["ytd_ach"],
            "Current Run Rate": cell["current_rr"],
            "Current March Projection": cell["current_march"],
            "Scenario Achievement %": cell["milestone_pct"],
            "Scenario Target Amount": cell["march_required"],
            "Scenario Required Run Rate": cell["scen_rr"],
            "Run Rate Uplift %": cell["rr_change_pct"],
            "Incremental Amount": cell["incremental_sales"],
        })
    overall = model.cell(sales)
    rows.append({
        "Segment": "Overall",
        "FY Target": overall["fy_target"],
        "YTD Achievement": overall["ytd_ach"],
        "Current Run Rate": overall["current_rr"],
        "Current March Projection": overall["current_march"],
        "Scenario Achievement %": overall["march_pct"],
        "Scenario Target Amount": overall["march_required"],
        "Scenario Required Run Rate": overall["scen_rr"],
        "Run Rate Uplift %": overall["rr_change_pct"],
        "Incremental Amount": overall["incremental_sales"],
    })
    formats = {
        "Segment": "txt", "FY Target": "cr", "YTD Achievement": "cr",
        "Current Run Rate": "cr", "Current March Projection": "cr",
        "Scenario Achievement %": "pct", "Scenario Target Amount": "cr",
        "Scenario Required Run Rate": "cr", "Run Rate Uplift %": "pct_signed",
        "Incremental Amount": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_momentum_analysis(cell: Dict[str, Any]) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 - month-by-month momentum trajectory."""
    trajectory = cell.get("trajectory") or []
    current_rr = _num(cell.get("current_rr"))
    growth = cell.get("momentum_g")
    leakage = _z(cell.get("leakage"))
    fy_target = cell.get("fy_target")
    cumulative = _z(cell.get("ytd_ach"))
 
    rows = []
    previous = current_rr
    for position, month in enumerate(FUTURE_MONTHS):
        run_rate = trajectory[position] if position < len(trajectory) else None
        cumulative += _z(run_rate)
        if position < MONTHS_JUL_JAN:
            mom = growth if growth is not None else safe_div(run_rate, previous)
            if growth is None and mom is not None:
                mom = mom - 1.0
            phase = "Momentum build-up"
        else:
            mom = -leakage
            phase = "Feb-Mar leakage"
        rows.append({
            "Month": month,
            "Phase": phase,
            "Current Run Rate": current_rr,
            "Required Scenario Run Rate": run_rate,
            "MoM Growth": mom,
            "Cumulative Achievement": cumulative,
            "Achievement %": safe_div(cumulative, fy_target),
        })
        previous = run_rate
    formats = {
        "Month": "txt", "Phase": "txt", "Current Run Rate": "cr",
        "Required Scenario Run Rate": "cr", "MoM Growth": "pct_signed",
        "Cumulative Achievement": "cr", "Achievement %": "pct",
    }
    return pd.DataFrame(rows), formats
 
 
def build_momentum_by_group(
    model: ScenarioModel, sales: str, dimension: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 applied independently to asset classes or to verticals."""
    if dimension == "asset":
        keys = [("Asset", asset, {"asset": asset}) for asset in ASSETS]
        first_column = "Asset"
    else:
        keys = [("Vertical", v, {"vertical": v}) for v in model.available_verticals()]
        first_column = "Vertical"
 
    rows = []
    for _, name, filters in keys:
        cell = model.cell(sales, **filters)
        rows.append({
            first_column: name,
            "Current Run Rate": cell["current_rr"],
            "Required MoM Momentum": cell["momentum_g"],
            "January Achievement": cell["jan_amount"],
            "January Achievement %": cell["jan_pct"],
            "Feb-Mar Leakage": cell.get("leakage"),
            "March Achievement": cell["march_amount"],
            "March Achievement %": cell["march_pct"],
            "Headroom / Shortfall": cell["headroom_amt"],
        })
    formats = {
        first_column: "txt", "Current Run Rate": "cr",
        "Required MoM Momentum": "pct_signed", "January Achievement": "cr",
        "January Achievement %": "pct", "Feb-Mar Leakage": "pct",
        "March Achievement": "cr", "March Achievement %": "pct",
        "Headroom / Shortfall": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_monthly_revenue(model: ScenarioModel, basis: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 - monthly revenue implied by the momentum trajectory."""
    trajectories = {
        asset: (model.cell(basis, asset=asset).get("trajectory") or [])
        for asset in ASSETS
    }
    rows = []
    for position, month in enumerate(FUTURE_MONTHS):
        row = {"Month": month}
        total = 0.0
        for asset in ASSETS:
            series = trajectories[asset]
            sales_amount = series[position] if position < len(series) else 0.0
            revenue = _z(sales_amount) * REVENUE_RATE[asset]
            row[f"{asset} Revenue"] = revenue
            total += revenue
        row["Total Revenue"] = total
        rows.append(row)
    formats = {"Month": "txt", "Total Revenue": "cr1"}
    for asset in ASSETS:
        formats[f"{asset} Revenue"] = "cr1"
    return pd.DataFrame(rows), formats
 
 
def build_leakage_sensitivity(
    model: ScenarioModel, basis: str
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Scenario 7 - how the March outcome moves with the leakage assumption."""
    subset = filter_grid(model.grid, sales=basis)
    fy_target = subset["fy_target"].sum()
    ytd_target = subset["ytd_target"].sum()
    ytd_ach = subset["ytd_ach"].sum()
    rows = []
    for leakage in (0.0, 0.10, 0.20, 0.30):
        params = dict(model.params)
        params["leakage"] = leakage
        cell = calculate_scenario_7(fy_target, ytd_target, ytd_ach, params)
        rows.append({
            "Feb-Mar Leakage": leakage,
            "Required MoM Momentum": cell["momentum_g"],
            "January Achievement %": cell["jan_pct"],
            "March Achievement %": cell["march_pct"],
            "Headroom / Shortfall": cell["headroom_amt"],
        })
    formats = {
        "Feb-Mar Leakage": "pct", "Required MoM Momentum": "pct_signed",
        "January Achievement %": "pct", "March Achievement %": "pct",
        "Headroom / Shortfall": "cr_signed",
    }
    return pd.DataFrame(rows), formats
 
 
def build_scenario_guide(model: ScenarioModel, basis: str) -> pd.DataFrame:
    rows = []
    for scenario_id in SCENARIO_ORDER:
        meta = SCENARIOS[scenario_id]
        rows.append({
            "Scenario": meta["label"],
            "Description": meta["explanation"],
            "Milestone": meta["milestone"],
            "Selected": "Yes" if scenario_id == model.scenario_id else "",
        })
    rows.append({
        "Scenario": "Revenue methodology",
        "Description": (
            "Revenue is estimated at asset-class level using 60 bps for Equity, 20 bps for Debt "
            f"and 10 bps for Liquid, applied to {SALES_LABEL[basis]} only so that Gross Sales and "
            "Net Sales revenue are never double counted."
        ),
        "Milestone": "",
        "Selected": "",
    })
    rows.append({
        "Scenario": "Timeline assumption",
        "Description": (
            "April, May and June are complete. Three months completed, nine months remaining "
            "(July-January is seven months, February-March is two months). The current run rate "
            "is YTD achievement divided by three."
        ),
        "Milestone": "",
        "Selected": "",
    })
    if model.scenario_id == 3:
        rows.append({
            "Scenario": "Scenario 3 setting",
            "Description": f"Feb-Mar run-rate dip: {fmt_pct(model.params.get('dip', S3_DEFAULT_DIP))}",
            "Milestone": "", "Selected": "",
        })
    if model.scenario_id == 7:
        rows.append({
            "Scenario": "Scenario 7 settings",
            "Description": (
                f"January target: {fmt_pct(model.params.get('jan_target', S7_DEFAULT_JAN_TARGET))} · "
                f"March target: {fmt_pct(model.params.get('mar_target', S7_DEFAULT_MAR_TARGET))} · "
                f"Feb-Mar leakage: {fmt_pct(model.params.get('leakage', S7_DEFAULT_LEAKAGE))}"
            ),
            "Milestone": "", "Selected": "",
        })
    return pd.DataFrame(rows)
 
 
# =============================================================================
# 11. EXCEL EXPORT
# =============================================================================
 
def _round_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for column in out.columns:
        if pd.api.types.is_numeric_dtype(out[column]):
            out[column] = out[column].astype(float).round(4)
    return out
 
 
def make_export_excel(model: ScenarioModel, basis: str) -> bytes:
    """Build the management export workbook for the selected scenario."""
    sheets: List[Tuple[str, pd.DataFrame]] = []
    sheets.append(("Scenario Guide", build_scenario_guide(model, basis)))
    sheets.append(("Current Baseline", build_current_overview(model)[0]))
    sheets.append(("Current vs Scenario", build_comparison(model)[0]))
    sheets.append(("Revenue Impact", build_revenue_impact(model, basis)[0]))
    sheets.append(("Retail-DHNI-VRM Summary", build_vertical_summary(model)[0]))
    sheets.append(("Gross Sales Breakdown", build_asset_breakdown(model, "GS")[0]))
    sheets.append(("Net Sales Breakdown", build_asset_breakdown(model, "NS")[0]))
 
    segment_model = model if model.scenario_id == 6 else ScenarioModel(6, model.grid, model.params)
    segment_frames = []
    for sales in SALES_TYPES:
        frame = build_segment_scenario_analysis(segment_model, sales)[0]
        frame.insert(0, "Sales", SALES_LABEL[sales])
        segment_frames.append(frame)
    sheets.append(("Scenario 6 Segments", pd.concat(segment_frames, ignore_index=True)))
 
    momentum_model = model if model.scenario_id == 7 else ScenarioModel(7, model.grid, model.params)
    momentum_frames = []
    for sales in SALES_TYPES:
        overall = momentum_model.cell(sales)
        frame = build_momentum_analysis(overall)[0]
        frame.insert(0, "Sales", SALES_LABEL[sales])
        momentum_frames.append(frame)
    for sales in SALES_TYPES:
        for dimension in ("asset", "vertical"):
            frame = build_momentum_by_group(momentum_model, sales, dimension)[0]
            frame.insert(0, "Sales", SALES_LABEL[sales])
            momentum_frames.append(frame)
    sheets.append(("Scenario 7 Momentum", pd.concat(momentum_frames, ignore_index=True)))
    sheets.append(("S7 Monthly Revenue", build_monthly_revenue(momentum_model, basis)[0]))
 
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets:
            _round_frame(frame).to_excel(writer, sheet_name=name[:31], index=False)
        _style_workbook(writer)
    return buffer.getvalue()
 
 
def _style_workbook(writer: Any) -> None:
    """Bold headers, frozen top row and sensible column widths."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:  # pragma: no cover - openpyxl always present in the stack
        return
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            longest = 0
            letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                longest = max(longest, min(len(value), 60))
            worksheet.column_dimensions[letter].width = max(12, min(longest + 3, 62))
 
 
# =============================================================================
# 12. PRESENTATION LAYER - THEME
# =============================================================================
 
CUSTOM_CSS = """
<style>
:root {
    --ivory: #F4F7FB;
    --ivory-card: rgba(255,255,255,.68);
    --ivory-soft: #EAF0F8;
    --gold: #356AE6;
    --gold-deep: #4B5FA5;
    --gold-soft: rgba(76,99,133,.22);
    --ink: #172033;
    --muted: #667085;
    --green: #159B73;
    --red: #D85C67;
}
 
.stApp, [data-testid="stAppViewContainer"] {
    background: radial-gradient(circle at 8% 8%, rgba(53,106,230,.10), transparent 28%), radial-gradient(circle at 92% 10%, rgba(108,92,231,.09), transparent 25%), linear-gradient(180deg, #F8FAFD 0%, var(--ivory) 100%);
}
[data-testid="stHeader"] { background: transparent; }
 
.block-container, [data-testid="stMainBlockContainer"] {
    padding-top: 1.25rem;
    padding-bottom: 3rem;
    max-width: 1680px;
}
 
.stApp, .stApp p, .stApp span, .stApp li, .stApp label,
.stApp h1, .stApp h2, .stApp h3, .stApp h4 {
    color: var(--ink);
}
 
.app-title {
    font-size: 1.8rem;
    font-weight: 800;
    color: var(--ink);
    margin: 0 0 3px 0;
    letter-spacing: -0.02em;
}
.app-sub {
    color: var(--muted);
    font-size: 0.9rem;
    margin-bottom: 4px;
}
.app-rule {
    height: 1px;
    background: linear-gradient(90deg, var(--gold), transparent);
    margin: 12px 0 5px 0;
}
 
.section-label {
    font-size: 0.74rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: var(--gold-deep);
    font-weight: 800;
    margin: 22px 0 10px 0;
    border-left: 4px solid var(--gold);
    padding: 7px 0 7px 11px;
    background: linear-gradient(90deg, rgba(184,146,59,.10), transparent 55%);
    border-radius: 0 8px 8px 0;
}
 
.subsection-title {
    color: var(--gold-deep);
    font-size: 0.76rem;
    font-weight: 800;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin: 13px 0 5px 0;
}
 
.note {
    color: var(--muted);
    font-size: 0.79rem;
    line-height: 1.5;
    margin: 5px 0 12px 0;
}
 
.callout, .scenario-highlight {
    background: var(--ivory-card);
    border: 1px solid var(--gold-soft);
    border-left: 4px solid var(--gold);
    border-radius: 16px;
    padding: 12px 15px;
    color: var(--ink);
    font-size: 0.87rem;
    line-height: 1.55;
    margin: 7px 0 12px 0;
    box-shadow: 0 3px 12px rgba(31,45,68,.06);
}
.callout-warn { border-left-color: #C27A28; }
.callout-ok { border-left-color: var(--green); }
.tag-ok { color: var(--green); font-weight: 700; }
.tag-warn { color: #A96A22; font-weight: 700; }
 
[data-testid="stMetric"], [data-testid="metric-container"] {
    background: linear-gradient(145deg, rgba(255,255,255,.82) 0%, rgba(235,242,251,.68) 100%);
    border: 1px solid rgba(53,106,230,.22);
    border-top: 2px solid rgba(53,106,230,.55);
    border-radius: 16px;
    padding: 13px 15px 12px 15px;
    min-height: 108px;
    box-shadow: 0 10px 30px rgba(31,45,68,.07);
}
[data-testid="stMetricLabel"], [data-testid="stMetricLabel"] p,
[data-testid="stMetricLabel"] div {
    color: var(--muted) !important;
    font-size: 0.70rem !important;
    letter-spacing: 0.07em;
    text-transform: uppercase;
    font-weight: 700;
}
[data-testid="stMetricValue"] {
    color: var(--ink) !important;
    font-size: 1.43rem !important;
    font-weight: 800;
}
[data-testid="stMetricDelta"] { font-size: 0.76rem !important; }
 
[data-testid="stDataFrame"] {
    border: 1px solid var(--gold-soft);
    border-radius: 16px;
    overflow: hidden;
    background: var(--ivory-card);
    box-shadow: 0 2px 8px rgba(31,45,68,.05);
}
 
.stTabs [data-baseweb="tab-list"] {
    gap: 5px;
    border-bottom: 1px solid var(--gold-soft);
}
.stTabs [data-baseweb="tab"] {
    color: var(--muted);
    background: rgba(255,255,255,.34);
    border-radius: 8px 8px 0 0;
}
.stTabs [aria-selected="true"] {
    color: var(--gold-deep) !important;
    font-weight: 800;
    border-bottom: 2px solid var(--accent);
}
 
[data-testid="stExpander"] {
    border: 1px solid var(--gold-soft) !important;
    border-radius: 16px;
    background: rgba(255,255,255,.54);
}
 
[data-testid="stSidebar"], [data-testid="stSidebarContent"] {
    background: rgba(239,244,251,.76) !important;
    border-right: 1px solid var(--gold-soft);
}
[data-testid="stSidebar"] * { color: var(--ink) !important; }
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p,
[data-testid="stSidebar"] small { color: var(--muted) !important; }
 
[data-testid="stSidebar"] [data-baseweb="select"] > div,
[data-testid="stSidebar"] [data-baseweb="input"],
[data-testid="stSidebar"] input,
[data-baseweb="select"] > div {
    background: var(--ivory-card) !important;
    border-color: var(--gold-soft) !important;
    color: var(--ink) !important;
    border-radius: 9px !important;
}
 
[data-testid="stSidebar"] [data-testid="stExpander"] {
    background: rgba(255,255,255,.58) !important;
    border: 1px solid var(--gold-soft) !important;
}
 
.sidebar-title {
    font-size: 0.74rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    font-weight: 800;
    color: var(--gold-deep);
    margin-bottom: 7px;
}
.sidebar-card {
    background: var(--ivory-card);
    border: 1px solid var(--gold-soft);
    border-left: 4px solid var(--gold);
    border-radius: 16px;
    padding: 11px 12px;
    margin: 5px 0 13px 0;
    box-shadow: 0 3px 10px rgba(31,45,68,.05);
}
.sidebar-card .s-name {
    font-weight: 800;
    font-size: 0.88rem;
    color: var(--ink);
}
.sidebar-card .s-body {
    font-size: 0.78rem;
    color: var(--muted);
    line-height: 1.5;
    margin-top: 4px;
}
.sidebar-card .s-milestone {
    font-size: 0.75rem;
    color: var(--gold-deep);
    margin-top: 7px;
    font-weight: 700;
}
 
.stButton > button,
.stDownloadButton > button {
    background: var(--gold-deep) !important;
    color: #FFFDF7 !important;
    border: 1px solid var(--gold-deep) !important;
    border-radius: 9px !important;
    font-weight: 700 !important;
}
.stButton > button:hover,
.stDownloadButton > button:hover {
    background: var(--gold) !important;
    color: #FFFDF7 !important;
    border-color: var(--gold) !important;
}
 
[data-testid="stVegaLiteChart"] text { fill: var(--ink) !important; }
[data-testid="stVegaLiteChart"] .role-axis-grid line { stroke: #DCE4F0 !important; }
[data-testid="stVegaLiteChart"] .role-axis line,
[data-testid="stVegaLiteChart"] .role-axis path { stroke: var(--gold-soft) !important; }
 
hr { border-color: var(--gold-soft) !important; }
 
.gold-star-card {
    background: linear-gradient(145deg, #356AE6 0%, #6257D9 100%);
    border: 1px solid rgba(255,255,255,.22);
    border-radius: 14px;
    padding: 16px 15px;
    min-height: 138px;
    box-shadow: 0 7px 18px rgba(53,106,230,.20);
    color: #FFFFFF;
}
.gold-star-rank {
    font-size: .72rem;
    font-weight: 900;
    letter-spacing: .12em;
    text-transform: uppercase;
    color: rgba(255,255,255,.72);
}
.gold-star-name {
    margin-top: 7px;
    font-size: 1.05rem;
    font-weight: 900;
    color: #FFFFFF;
}
.gold-star-stat {
    margin-top: 9px;
    font-size: .82rem;
    line-height: 1.55;
    color: rgba(255,255,255,.84);
}
.category-pill {
    display: inline-block;
    background: rgba(53,106,230,.08);
    border: 1px solid rgba(53,106,230,.22);
    color: #3159AE;
    padding: 4px 9px;
    border-radius: 999px;
    font-size: .72rem;
    font-weight: 800;
    margin: 2px 4px 2px 0;
}
.final-output-box {
    background: linear-gradient(135deg, rgba(24,167,184,.13) 0%, rgba(53,106,230,.13) 100%);
    border: 1px solid rgba(53,106,230,.20);
    border-radius: 13px;
    padding: 13px 15px;
    color: #172033;
    box-shadow: 0 6px 16px rgba(31,45,68,.06);
}
 
/* Apple-inspired management cockpit: clean hierarchy, glass surfaces, restrained motion. */
body, .stApp { font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display", "SF Pro Text", Inter, system-ui, sans-serif !important; }
.stButton > button, .stDownloadButton > button, .stSelectbox, .stMultiSelect, .stSlider { border-radius: 14px !important; }
.stButton > button, .stDownloadButton > button { transition: transform .18s ease, box-shadow .18s ease, background .18s ease; box-shadow: 0 2px 10px rgba(0,0,0,.06); }
.stButton > button:hover, .stDownloadButton > button:hover { transform: translateY(-1px); box-shadow: 0 8px 20px rgba(0,0,0,.10); }
[data-testid="stMetric"], [data-testid="metric-container"], .sidebar-card, .callout, .scenario-highlight { backdrop-filter: blur(20px) saturate(1.2); -webkit-backdrop-filter: blur(20px) saturate(1.2); }
[data-testid="stMetric"], [data-testid="metric-container"] { border-radius: 18px !important; border-top-width: 2px !important; box-shadow: 0 8px 28px rgba(50,40,20,.08) !important; }
.stSlider [role="slider"] { transition: transform .15s ease; }
.stSlider [role="slider"]:active { transform: scale(1.08); }
.channel-simulator-note { padding: 12px 16px; border-radius: 16px; background: rgba(255,255,255,.60); border: 1px solid rgba(53,106,230,.16); color: var(--muted); margin-bottom: 12px; }
@media (prefers-reduced-motion: reduce) { * { animation-duration: .01ms !important; transition-duration: .01ms !important; } }
 
</style>
"""
 
 
def rerun() -> None:
    handler = getattr(st, "rerun", None) or getattr(st, "experimental_rerun", None)
    if handler is not None:
        handler()
 
 
def _dataframe_kwargs() -> Dict[str, Any]:
    try:
        parameters = inspect.signature(st.dataframe).parameters
    except (TypeError, ValueError):
        return {}
    if "width" in parameters and parameters["width"].default == "stretch":
        return {}
    if "use_container_width" in parameters:
        return {"use_container_width": True}
    return {}
 
 
def show_table(frame: pd.DataFrame, formats: Optional[Dict[str, str]] = None) -> None:
    display = format_table(frame, formats) if formats else frame
    try:
        st.dataframe(display, hide_index=True, **_dataframe_kwargs())
    except TypeError:  # pragma: no cover - very old Streamlit
        st.dataframe(display)
 
 
def section(title: str) -> None:
    st.markdown(f"<div class='section-label'>{title}</div>", unsafe_allow_html=True)
 
 
def callout(text: str, tone: str = "") -> None:
    css = "callout" + (f" callout-{tone}" if tone else "")
    st.markdown(f"<div class='{css}'>{text}</div>", unsafe_allow_html=True)
 
 
def kpi_row(items: Sequence[Tuple]) -> None:
    if not items:
        return
    columns = st.columns(len(items))
    for column, item in zip(columns, items):
        label, value = item[0], item[1]
        delta = item[2] if len(item) > 2 else None
        delta_color = item[3] if len(item) > 3 else "normal"
        with column:
            if delta is None:
                st.metric(label, value)
            else:
                st.metric(label, value, delta, delta_color=delta_color)
 
 
# =============================================================================
# 13. UPLOAD GATE
# =============================================================================
 
def render_upload_screen() -> None:
    _, middle, _ = st.columns([1, 2, 1])
    with middle:
        _render_upload_body()
 
 
def _render_upload_body() -> None:
    st.markdown(f"<div class='app-title'>{APP_TITLE}</div>", unsafe_allow_html=True)
    st.markdown(
        "<div class='app-sub'>Upload the RM scorecard workbook to open the dashboard. "
        "Nothing is stored beyond this session.</div>",
        unsafe_allow_html=True,
    )
    st.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
 
    uploaded = st.file_uploader(
        "Workbook", type=["xlsx", "xlsm"], label_visibility="collapsed",
    )
    if uploaded is not None:
        payload = uploaded.getvalue()
        try:
            load_workbook(payload)
            parse_final_dashboard_metrics(payload)
        except WorkbookError as error:
            st.error(str(error))
        except Exception:  # pragma: no cover - defensive, never show a traceback
            st.error(
                "The workbook could not be read. Please check that it is a valid Excel file "
                "containing RM Retail Sales, RM DHNI, VRM and FINAL."
            )
        else:
            st.session_state["workbook"] = payload
            rerun()
 
 
# =============================================================================
# 14. SIDEBAR
# =============================================================================
 
def render_channel_controls(records: pd.DataFrame) -> Dict[str, Any]:
    """Apple-style channel mapping controls for Scenario 8/9."""
    suggestions = identify_channels(records)
    mapping: Dict[str, Any] = dict(st.session_state.get("channel_mapping") or suggestions)
    usable_columns = [f for f in META_FIELDS + ["Vertical"] if f in records.columns and text_column(records, f).ne("").any()]
    with st.sidebar.expander("Channel mapping", expanded=False):
        st.caption("Map workbook metadata to Digital, VRM, EM, B30, T30, T8, DHNI, Retail and Institutional. Unmapped rows remain Unclassified.")
        for channel in CHANNELS:
            options = ["(not mapped)"] + usable_columns
            current = mapping.get(channel, {}).get("column", "(not mapped)")
            idx = options.index(current) if current in options else 0
            column = st.selectbox(channel, options, index=idx, key=f"ch_col_{channel}")
            if column == "(not mapped)":
                mapping.pop(channel, None); continue
            values = sorted({v for v in text_column(records, column) if v.strip()})
            preset = [v for v in mapping.get(channel, {}).get("values", []) if v in values]
            chosen = st.multiselect(f"{channel} values", values, default=preset, key=f"ch_vals_{channel}_{column}")
            if chosen: mapping[channel] = {"column": column, "values": list(chosen)}
            else: mapping.pop(channel, None)
    st.session_state["channel_mapping"] = mapping
    return mapping
 
 
def render_sidebar(records: pd.DataFrame) -> Tuple[int, Dict[str, Any], str, Dict[str, Any]]:
    sidebar = st.sidebar
    sidebar.markdown("<div class='sidebar-title'>Scenario Navigator</div>", unsafe_allow_html=True)
 
    scenario_options = SCENARIO_ORDER
    scenario_id = sidebar.selectbox(
        "Scenario",
        scenario_options,
        index=0,
        key="scenario_selector_dropdown",
        format_func=lambda sid: (
            f"Scenario {sid} · {SCENARIOS[sid]['name']}"
        ),
        label_visibility="collapsed",
    )
    meta = SCENARIOS[scenario_id]
    sidebar.markdown(
        "<div class='sidebar-card'>"
        f"<div class='s-name'>Scenario {scenario_id} · {meta['name']}</div>"
        f"<div class='s-body'>{meta['explanation']}</div>"
        f"<div class='s-milestone'>Target milestone: {meta['milestone']}</div>"
        "</div>",
        unsafe_allow_html=True,
    )
 
    params: Dict[str, Any] = {
        "dip": S3_DEFAULT_DIP,
        "jan_target": S7_DEFAULT_JAN_TARGET,
        "mar_target": S7_DEFAULT_MAR_TARGET,
        "leakage": S7_DEFAULT_LEAKAGE,
        "channel_growth": dict(S8_DEFAULT_GROWTH),
        "channel_jan_target": dict(S8_DEFAULT_JAN_TARGET),
        "channel_mar_target": dict(S8_DEFAULT_MAR_TARGET),
    }
 
    if scenario_id == 3:
        sidebar.markdown("<div class='sidebar-title'>Scenario controls</div>", unsafe_allow_html=True)
        params["dip"] = sidebar.slider(
            "Feb-Mar Run-Rate Dip", min_value=0, max_value=60,
            value=int(S3_DEFAULT_DIP * 100), step=5, format="%d%%", key="s3_dip",
        ) / 100.0
 
    if scenario_id == 7:
        sidebar.markdown("<div class='sidebar-title'>Scenario controls</div>", unsafe_allow_html=True)
        params["jan_target"] = sidebar.slider(
            "January Achievement Target", min_value=90, max_value=120,
            value=int(S7_DEFAULT_JAN_TARGET * 100), step=1, format="%d%%", key="s7_jan",
        ) / 100.0
        params["mar_target"] = sidebar.slider(
            "March Achievement Target", min_value=90, max_value=120,
            value=int(S7_DEFAULT_MAR_TARGET * 100), step=1, format="%d%%", key="s7_mar",
        ) / 100.0
        params["leakage"] = sidebar.slider(
            "Feb-Mar AUM Leakage", min_value=0, max_value=30,
            value=int(S7_DEFAULT_LEAKAGE * 100), step=1, format="%d%%", key="s7_leak",
        ) / 100.0
 
    if scenario_id in (8, 9):
        sidebar.markdown("<div class='sidebar-title'>Channel simulator controls</div>", unsafe_allow_html=True)
        params["leakage"] = sidebar.slider(
            "Feb-Mar AUM Leakage", 0, 30, int(S8_DEFAULT_LEAKAGE * 100), 1, format="%d%%", key="s8_leakage"
        ) / 100.0
        channel_mapping = render_channel_controls(records)
        if scenario_id == 8:
            with sidebar.expander("Channel assumptions", expanded=True):
                for channel in CHANNELS:
                    cols = st.columns(3)
                    params["channel_growth"][channel] = cols[0].slider(
                        f"{channel} · MoM", -20, 30, int(S8_DEFAULT_GROWTH[channel] * 100), 1, format="%d%%", key=f"s8_g_{channel}"
                    ) / 100.0
                    params["channel_jan_target"][channel] = cols[1].slider(
                        "Jan 2027", 80, 180, int(S8_DEFAULT_JAN_TARGET[channel] * 100), 1, format="%d%%", key=f"s8_j_{channel}"
                    ) / 100.0
                    params["channel_mar_target"][channel] = cols[2].slider(
                        "Mar 2027", 80, 200, int(S8_DEFAULT_MAR_TARGET[channel] * 100), 1, format="%d%%", key=f"s8_m_{channel}"
                    ) / 100.0
        else:
            params["optimizer_target"] = sidebar.slider("Portfolio March ambition", 100, 180, 120, 1, format="%d%%", key="s9_target") / 100.0
    else:
        channel_mapping = render_channel_controls(records) if scenario_id in (8, 9) else {}
 
    mapping = render_segment_controls(records)
    params["channel_mapping"] = channel_mapping
 
    with sidebar.expander("Assumptions", expanded=False):
        st.caption(
            "Revenue basis: Net Sales only. "
            "Revenue rates: Equity 60 bps · Debt 20 bps · Liquid 10 bps. "
            "Timeline: April-June complete, 9 months remaining, "
            "July-January 7 months, February-March 2 months."
        )
    basis = "NS"
 
    sidebar.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
    if sidebar.button("Use another workbook"):
        reset_workbook()
 
    return scenario_id, params, basis, mapping
 
 
def render_segment_controls(records: pd.DataFrame) -> Dict[str, Any]:
    """Configurable segment classification for Scenario 6."""
    suggestions = identify_segments(records)
    mapping: Dict[str, Any] = dict(st.session_state.get("segment_mapping") or suggestions)
 
    usable_columns = [
        field for field in META_FIELDS
        if field in records.columns and text_column(records, field).ne("").any()
    ]
 
    with st.sidebar.expander("Segment mapping (Scenario 6)", expanded=False):
        st.caption(
            "Digital, Retail B30 and Others are derived from workbook metadata. "
            "Adjust the classification here; every other record falls into Others."
        )
        for segment in ("Digital", "Retail B30"):
            options = ["(not mapped)"] + usable_columns
            current = mapping.get(segment, {}).get("column", "(not mapped)")
            index = options.index(current) if current in options else 0
            column = st.selectbox(
                f"{segment} identified by", options, index=index, key=f"seg_col_{segment}",
            )
            if column == "(not mapped)":
                mapping.pop(segment, None)
                continue
            values = sorted({v for v in text_column(records, column) if v.strip()})
            preset = [v for v in mapping.get(segment, {}).get("values", []) if v in values]
            chosen = st.multiselect(
                f"{segment} values", values, default=preset,
                key=f"seg_vals_{segment}_{column}",
            )
            if chosen:
                mapping[segment] = {"column": column, "values": list(chosen)}
            else:
                mapping.pop(segment, None)
 
    st.session_state["segment_mapping"] = mapping
    return mapping
 
 
# =============================================================================
# 15. DASHBOARD SECTIONS
# =============================================================================
 
def render_baseline(model: ScenarioModel) -> None:
    section("Current Baseline")
    cards = []
    for sales in SALES_TYPES:
        base = model.baseline(sales)
        cards.append((
            f"{sales} Current Run Rate", fmt_cr(base["current_rr"]),
            f"{fmt_pct(base['fy_completed_pct'])} of FY target booked", "off",
        ))
    for sales in SALES_TYPES:
        base = model.baseline(sales)
        delta = None if base["ytd_ach_pct"] is None else fmt_pts(base["ytd_ach_pct"] - 1.0)
        cards.append((f"{sales} Target Achieved %", fmt_pct(base["ytd_ach_pct"]), delta))
    for sales in SALES_TYPES:
        base = model.baseline(sales)
        delta = None if base["current_march_pct"] is None else fmt_pts(base["current_march_pct"] - 1.0)
        cards.append((f"{sales} March Projection %", fmt_pct(base["current_march_pct"]), delta))
    kpi_row(cards)
    st.markdown(
        "<div class='note'>Baseline is the Apr-Jun run rate carried forward for nine months. "
        "It never changes with the selected scenario.</div>",
        unsafe_allow_html=True,
    )
 
 
def scenario_cards(model: ScenarioModel, basis: str) -> List[Tuple]:
    kind = model.meta["kind"]
    cards: List[Tuple] = []
 
    if model.scenario_id == 6:
        for segment in model.available_segments():
            cell = model.cell(basis, segment=segment)
            delta = (
                None if cell["march_pct"] is None or cell["current_march_pct"] is None
                else fmt_pts(cell["march_pct"] - cell["current_march_pct"])
            )
            cards.append((f"{segment} Achievement", fmt_pct(cell["march_pct"]), delta))
        overall = model.cell(basis)
        cards.append((
            "Overall Achievement", fmt_pct(overall["march_pct"]),
            None if overall["march_pct"] is None or overall["current_march_pct"] is None
            else fmt_pts(overall["march_pct"] - overall["current_march_pct"]),
        ))
        cards.append((
            "Scenario March Projection", fmt_cr(overall["march_amount"]),
            fmt_cr_signed(overall["incremental_sales"]),
        ))
        return cards
 
    if model.scenario_id == 7:
        cell = model.cell(basis)
        momentum = cell["momentum_g"]
        cards.append((
            "Required MoM Momentum",
            fmt_pct(momentum) if momentum is not None else NA_TEXT,
            f"binding: {cell['binding']} milestone" if cell.get("binding") else None, "off",
        ))
        cards.append((
            "January Achievement", fmt_pct(cell["jan_pct"]),
            fmt_cr_signed(cell["jan_buffer"]) + " buffer" if cell["jan_buffer"] is not None else None,
        ))
        cards.append(("Feb-Mar Leakage", fmt_pct(cell.get("leakage")), None))
        cards.append((
            "March Achievement", fmt_pct(cell["march_pct"]),
            None if cell["march_pct"] is None or cell["current_march_pct"] is None
            else fmt_pts(cell["march_pct"] - cell["current_march_pct"]),
        ))
        cards.append((
            "March Headroom / Shortfall", fmt_cr_signed(cell["headroom_amt"]),
            fmt_pts(cell["headroom_pct"]) if cell["headroom_pct"] is not None else None,
        ))
        cards.append((
            "January Exit Run Rate", fmt_cr(cell["scen_rr"]),
            fmt_pct_signed(cell["rr_change_pct"]) if cell["rr_change_pct"] is not None else None,
        ))
        return cards
 
    for sales in SALES_TYPES:
        cell = model.cell(sales)
        label = "Scenario Run Rate" if kind == "runrate" else "Required Run Rate"
        cards.append((
            f"{sales} {label}", fmt_cr(cell["scen_rr"]),
            fmt_pct_signed(cell["rr_change_pct"]) if cell["rr_change_pct"] is not None else None,
        ))
 
    if kind == "jan_target":
        for sales in SALES_TYPES:
            cell = model.cell(sales)
            cards.append((
                f"{sales} January Achievement", fmt_pct(cell["jan_pct"]),
                fmt_cr(cell["jan_amount"]), "off",
            ))
    for sales in SALES_TYPES:
        cell = model.cell(sales)
        delta = (
            None if cell["march_pct"] is None or cell["current_march_pct"] is None
            else fmt_pts(cell["march_pct"] - cell["current_march_pct"])
        )
        cards.append((f"{sales} March Achievement", fmt_pct(cell["march_pct"]), delta))
    return cards
 
 
def render_comparison(model: ScenarioModel, basis: str) -> None:
    section(f"Current vs Scenario {model.scenario_id} · {model.meta['name']}")
    cards = scenario_cards(model, basis)
    if len(cards) <= 6:
        kpi_row(cards)
    else:
        kpi_row(cards[:4])
        kpi_row(cards[4:8])
 
    if model.scenario_id in (2, 5):
        implied = model.implied_milestones(basis)
        st.markdown(
            "<div class='note'>Equity carries the fixed ambition; Debt and Liquid share the "
            f"residual requirement in FY-target proportion — implied milestone of "
            f"{fmt_pct(implied['Debt'])} of FY target for both ({SALES_LABEL[basis]}).</div>",
            unsafe_allow_html=True,
        )
    if model.scenario_id == 3:
        st.markdown(
            f"<div class='note'>February-March run rate is set at "
            f"{fmt_pct(1 - model.params['dip'])} of the July-January required run rate.</div>",
            unsafe_allow_html=True,
        )
    if model.scenario_id == 2:
        st.markdown(
            "<div class='note'>February-March is assumed to continue at the July-January "
            "required run rate.</div>",
            unsafe_allow_html=True,
        )
 
 
def render_revenue_kpis(model: ScenarioModel, basis: str) -> Dict[str, Any]:
    section("Revenue / Earnings Impact")
    bundle = revenue_bundle(model, basis)
    baseline_total = bundle["baseline"]["total"]
    scenario_total = bundle["scenario"]["total"]
    incremental = bundle["incremental"]
    kpi_row([
        ("Baseline Revenue", fmt_cr(baseline_total, 1), "current run rate", "off"),
        ("Scenario Revenue", fmt_cr(scenario_total, 1), fmt_cr_signed(incremental["total"], 1)),
        ("Incremental Revenue", fmt_cr_signed(incremental["total"], 1),
         fmt_pct_signed(incremental["uplift_pct"]) if incremental["uplift_pct"] is not None else None),
        ("Revenue Uplift %",
         fmt_pct(incremental["uplift_pct"]) if incremental["uplift_pct"] is not None else NA_TEXT,
         f"on {SALES_LABEL[basis]}", "off"),
    ])
    st.markdown(
        f"<div class='note'>Revenue is calculated at asset-class level (Equity 60 bps, Debt 20 bps, "
        f"Liquid 10 bps) on {SALES_LABEL[basis]} only — no blended rate, no double counting.</div>",
        unsafe_allow_html=True,
    )
    return bundle
 
 
def render_detail_expander(model: ScenarioModel) -> None:
    with st.expander("Detailed baseline and comparison numbers", expanded=False):
        st.markdown("**Current baseline**")
        frame, formats = build_current_overview(model)
        show_table(frame, formats)
        st.markdown("**Current vs selected scenario**")
        frame, formats = build_comparison(model)
        show_table(frame, formats)
 
 
def render_vertical_section(model: ScenarioModel) -> None:
    section("Channel Drill-Down")
 
    available = model.available_verticals()
    if not available:
        st.info("No Retail / DHNI / VRM channel data is available.")
        return
 
    selected_vertical = st.selectbox(
        "Select Channel",
        available,
        index=0,
        key="channel_drilldown",
        help="Choose Retail, DHNI or VRM to inspect the selected scenario.",
    )
 
    frame, formats = build_vertical_summary(model)
    filtered = frame.loc[frame["Vertical"] == selected_vertical].copy()
 
    tabs = st.tabs([SALES_LABEL["GS"], SALES_LABEL["NS"]])
    for tab, sales in zip(tabs, SALES_TYPES):
        with tab:
            sales_frame = filtered.loc[filtered["Sales"] == SALES_LABEL[sales]].copy()
            if not sales_frame.empty:
                sales_frame = sales_frame.drop(columns=["Vertical", "Sales"])
                show_table(sales_frame, formats)
            else:
                st.info(f"No {SALES_LABEL[sales]} data for {selected_vertical}.")
 
    st.markdown(
        "<div class='note'>Choose one channel at a time for a cleaner management view. "
        "Gross Sales and Net Sales remain alternative revenue bases and are never added together.</div>",
        unsafe_allow_html=True,
    )
 
 
def render_asset_section(model: ScenarioModel) -> None:
    section("Asset Class Drill-Down")
 
    selected_asset = st.selectbox(
        "Select Asset Class",
        ASSETS,
        index=0,
        key="asset_class_drilldown",
        help="Choose Equity, Debt or Liquid to compare that asset across channels.",
    )
 
    tabs = st.tabs([SALES_LABEL["GS"], SALES_LABEL["NS"]])
    for tab, sales in zip(tabs, SALES_TYPES):
        with tab:
            frame, formats = build_asset_breakdown(model, sales)
            filtered = frame.loc[frame["Asset"] == selected_asset].copy()
            if not filtered.empty:
                filtered = filtered.drop(columns=["Asset"])
                show_table(filtered, formats)
            else:
                st.info(f"No {SALES_LABEL[sales]} data for {selected_asset}.")
 
 
def render_segment_section(model: ScenarioModel, basis: str, counts: Dict[str, int]) -> None:
    section("Scenario 6 · Segment Analysis")
    unmapped = [s for s in ("Digital", "Retail B30") if counts.get(s, 0) == 0]
    if unmapped:
        missing = " and ".join(unmapped)
        callout(
            f"<b>Segment validation:</b> no records could be reliably classified as {missing} "
            "from the workbook metadata (MKT TYPE, Type, ZONE, REGION, EM City). "
            "Those records have <b>not</b> been reallocated — they remain in Others, and the "
            f"{missing} scenario uplift is therefore not applied. Use "
            "<i>Segment mapping (Scenario 6)</i> in the sidebar to point the classification at "
            "the correct column and values.",
            tone="warn",
        )
 
    present = " · ".join(
        f"{segment} {S6_SEGMENT_TARGETS[segment]:.0%} of FY target ({counts.get(segment, 0)} RMs)"
        for segment in SEGMENT_ORDER
    )
    st.markdown(f"<div class='note'>Scenario assumption — {present}.</div>", unsafe_allow_html=True)
 
    tabs = st.tabs([SALES_LABEL["GS"], SALES_LABEL["NS"]])
    for tab, sales in zip(tabs, SALES_TYPES):
        with tab:
            frame, formats = build_segment_scenario_analysis(model, sales)
            show_table(frame, formats)
 
    overall = model.cell(basis)
    lines = []
    for segment in model.available_segments():
        cell = model.cell(basis, segment=segment)
        uplift = (
            fmt_pct_signed(cell["rr_change_pct"]) if cell["rr_change_pct"] is not None else NA_TEXT
        )
        lines.append(
            f"<b>{segment}</b> moves from {fmt_pct(cell['current_march_pct'])} to "
            f"{fmt_pct(cell['march_pct'])} of FY target, needing {fmt_cr(cell['scen_rr'])} per month "
            f"({uplift} run-rate uplift)"
        )
    callout(
        f"On {SALES_LABEL[basis]}: " + "; ".join(lines) + ". Overall March achievement moves from "
        f"{fmt_pct(overall['current_march_pct'])} to {fmt_pct(overall['march_pct'])}, an improvement "
        f"of {fmt_cr_signed(overall['incremental_sales'])}."
    )
 
 
def render_momentum_section(model: ScenarioModel, basis: str) -> None:
    section("Scenario 7 · Momentum Analysis")
    cell = model.cell(basis)
    momentum = cell["momentum_g"]
 
    kpi_row([
        ("Current Run Rate", fmt_cr(cell["current_rr"]), f"Apr-Jun, {SALES_LABEL[basis]}", "off"),
        ("Required MoM Momentum", fmt_pct(momentum) if momentum is not None else NA_TEXT,
         f"binding: {cell['binding']}" if cell.get("binding") else None, "off"),
        ("January Target", fmt_cr(cell["jan_required"]),
         fmt_pct(model.params["jan_target"]) + " of FY target", "off"),
        ("January Scenario Achievement", fmt_cr(cell["jan_amount"]), fmt_pct(cell["jan_pct"]), "off"),
    ])
    kpi_row([
        ("January Buffer", fmt_cr_signed(cell["jan_buffer"]),
         fmt_pct_signed(cell["jan_buffer_pct"]) if cell["jan_buffer_pct"] is not None else None),
        ("Feb-Mar Leakage", fmt_pct(cell.get("leakage")),
         f"Feb {fmt_cr(cell['feb_mar_rr'])} · Mar {fmt_cr(cell.get('march_rr'))}", "off"),
        ("March Scenario Achievement", fmt_cr(cell["march_amount"]), fmt_pct(cell["march_pct"]), "off"),
        ("March Headroom / Shortfall", fmt_cr_signed(cell["headroom_amt"]),
         fmt_pts(cell["headroom_pct"]) if cell["headroom_pct"] is not None else None),
    ])
 
    if cell["feasible"]:
        callout(
            "<span class='tag-ok'>✓ TARGET ACHIEVABLE</span> — the momentum trajectory reaches the "
            f"January milestone and still clears the March ambition after "
            f"{fmt_pct(cell.get('leakage'))} Feb-Mar leakage.",
            tone="ok",
        )
    else:
        callout(
            "<span class='tag-warn'>⚠ ADDITIONAL MOMENTUM REQUIRED</span> — additional March sales "
            f"required: {fmt_cr(cell.get('additional_march_sales'))}; additional January run rate "
            f"required: {fmt_cr(cell.get('additional_jan_rr'))} per month.",
            tone="warn",
        )
    if cell.get("note"):
        st.markdown(f"<div class='note'>{cell['note']}</div>", unsafe_allow_html=True)
 
    momentum_text = fmt_pct(momentum) if momentum is not None else "a flat required"
    outcome = (
        f"{fmt_pct(cell['headroom_pct'])} headroom" if _z(cell["headroom_amt"]) >= 0
        else f"a {fmt_cr(abs(_z(cell['headroom_amt'])))} shortfall"
    )
    callout(
        f"<b>Momentum required:</b> the business needs to build approximately {momentum_text} "
        f"month-on-month momentum from July through January — lifting the monthly run rate from "
        f"{fmt_cr(cell['current_rr'])} to {fmt_cr(cell['scen_rr'])} by January — to reach the "
        f"{fmt_pct(model.params['jan_target'])} January milestone. With an assumed "
        f"{fmt_pct(cell.get('leakage'))} February-March leakage, the trajectory delivers "
        f"{fmt_pct(cell['march_pct'])} achievement by March against a "
        f"{fmt_pct(model.params['mar_target'])} ambition, creating {outcome}. "
        f"January buffer created before leakage: {fmt_cr_signed(cell['jan_buffer'])}."
    )
 
    st.markdown("**Monthly momentum trajectory**")
    frame, formats = build_momentum_analysis(cell)
    show_table(frame, formats)
 
    trajectory = cell.get("trajectory") or []
    if trajectory:
        chart = pd.DataFrame(
            {
                "Current run-rate trajectory": [_z(cell["current_rr"])] * len(trajectory),
                "Required momentum trajectory": trajectory,
            },
            index=MONTH_DATES[: len(trajectory)],
        )
        st.line_chart(chart)
        st.markdown(
            f"<div class='note'>January 2027 milestone: {fmt_cr(cell['jan_required'])} cumulative "
            f"({fmt_pct(model.params['jan_target'])} of FY target) — reached at a monthly run rate of "
            f"{fmt_cr(cell['scen_rr'])}, after which February and March step down by "
            f"{fmt_pct(cell.get('leakage'))} each.</div>",
            unsafe_allow_html=True,
        )
 
    tabs = st.tabs(["Asset class", "Retail / DHNI / VRM", "Leakage sensitivity", "Monthly revenue"])
    with tabs[0]:
        for sales in SALES_TYPES:
            st.markdown(f"**{SALES_LABEL[sales]}**")
            frame, formats = build_momentum_by_group(model, sales, "asset")
            show_table(frame, formats)
    with tabs[1]:
        for sales in SALES_TYPES:
            st.markdown(f"**{SALES_LABEL[sales]}**")
            frame, formats = build_momentum_by_group(model, sales, "vertical")
            show_table(frame, formats)
    with tabs[2]:
        frame, formats = build_leakage_sensitivity(model, basis)
        show_table(frame, formats)
        st.markdown(
            "<div class='note'>Momentum is re-solved at each leakage assumption, so the required "
            "July-January build changes with the February-March pressure.</div>",
            unsafe_allow_html=True,
        )
    with tabs[3]:
        frame, formats = build_monthly_revenue(model, basis)
        show_table(frame, formats)
        january_revenue = calculate_revenue(model.assets(basis), "jan_amount")
        march_revenue = calculate_revenue(model.assets(basis), "march_amount")
        baseline = calculate_baseline_revenue(model.assets(basis))
        st.markdown(
            f"<div class='note'>January scenario revenue {fmt_cr(january_revenue['total'], 1)} · "
            f"March scenario revenue {fmt_cr(march_revenue['total'], 1)} · baseline "
            f"{fmt_cr(baseline['total'], 1)} · incremental "
            f"{fmt_cr_signed(march_revenue['total'] - baseline['total'], 1)}.</div>",
            unsafe_allow_html=True,
        )
 
 
def render_revenue_detail(model: ScenarioModel, basis: str, bundle: Dict[str, Any]) -> None:
    section("Revenue by Asset Class")
    frame, formats = build_revenue_impact(model, basis)
    show_table(frame, formats)
 
    incremental = bundle["incremental"]
    parts = " + ".join(
        f"{asset} {fmt_cr_signed(incremental['by_asset'][asset], 1)}" for asset in ASSETS
    )
    contribution = " · ".join(
        f"{asset} {fmt_pct(incremental['contribution'][asset])}" for asset in ASSETS
    )
    callout(
        f"<b>Revenue bridge:</b> current run-rate revenue {fmt_cr(bundle['baseline']['total'], 1)} "
        f"+ {parts} = scenario revenue {fmt_cr(bundle['scenario']['total'], 1)}.<br>"
        f"<b>Scenario revenue contribution:</b> {contribution}."
    )
 
 
def render_export(model: ScenarioModel, basis: str) -> None:
    section("Export")
    try:
        payload = make_export_excel(model, basis)
    except Exception:  # pragma: no cover - defensive
        st.warning("The export could not be generated for the current selection.")
        return
    st.download_button(
        "Download Selected Scenario Analysis",
        data=payload,
        file_name=f"scenario_{model.scenario_id}_analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown(
        "<div class='note'>The workbook contains the scenario guide, baseline, comparison, revenue "
        "impact, Retail-DHNI-VRM summary, Gross and Net Sales breakdowns, Scenario 6 segment analysis, "
        "Scenario 7 momentum analysis and Scenario 7 monthly revenue.</div>",
        unsafe_allow_html=True,
    )
 
 
# =============================================================================
# 15A. RM PERFORMANCE SEGMENTATION PAGE
# =============================================================================
 
ACHIEVEMENT_BANDS: List[Tuple[str, float, Optional[float]]] = [
    ("100% and above", 1.00, None),
    ("90% - 100%", 0.90, 1.00),
    ("75% - 90%", 0.75, 0.90),
    ("50% - 75%", 0.50, 0.75),
    ("30% - 50%", 0.30, 0.50),
    ("Less than 30%", float("-inf"), 0.30),
]
ACHIEVEMENT_BAND_ORDER: List[str] = [item[0] for item in ACHIEVEMENT_BANDS]
 
 
def achievement_band(value: Any) -> str:
    """Map YTD-target achievement to the management bands requested for RMs."""
    ratio = _num(value)
    # User requested N/A to be treated/displayed as 0.
    ratio = 0.0 if ratio is None else ratio
 
    if ratio >= 1.00:
        return "100% and above"
    if ratio >= 0.90:
        return "90% - 100%"
    if ratio >= 0.75:
        return "75% - 90%"
    if ratio >= 0.50:
        return "50% - 75%"
    if ratio >= 0.30:
        return "30% - 50%"
    return "Less than 30%"
 
 
def _rm_identity_columns(records: pd.DataFrame) -> List[str]:
    preferred = [
        "Employee Name", "Emp Code", "ADID", "ZONE", "REGION",
        "EM City", "MKT TYPE", "Type", "Status",
    ]
    return [column for column in preferred if column in records.columns]
 
 
def build_rm_performance_detail(
    records: pd.DataFrame,
    vertical: str,
    sales: str,
    final_target: Optional[float] = None,
) -> pd.DataFrame:
    """
    Build one row per RM.
 
    Banding is based on overall YTD achievement / overall YTD target across
    Equity + Debt + Liquid. Run-rate projection annualises the first three
    completed months.
    """
    subset = records.loc[records["Vertical"] == vertical].copy()
    if subset.empty:
        return pd.DataFrame()
 
    identity = _rm_identity_columns(subset)
    out = subset[identity].copy()
 
    fy_cols = [f"{sales}_{asset}_fy" for asset in ASSETS]
    ytd_target_cols = [f"{sales}_{asset}_ytd_tgt" for asset in ASSETS]
    ach_cols = [f"{sales}_{asset}_ach" for asset in ASSETS]
 
    for columns in (fy_cols, ytd_target_cols, ach_cols):
        for column in columns:
            if column not in subset.columns:
                subset[column] = 0.0
 
    fy_matrix = subset[fy_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    ytd_target_matrix = subset[ytd_target_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    ach_matrix = subset[ach_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
 
    out["FY Target"] = fy_matrix.sum(axis=1)
    out["YTD Target"] = ytd_target_matrix.sum(axis=1)
    out["YTD Achievement"] = ach_matrix.sum(axis=1)
 
    out["YTD Achievement %"] = np.where(
        out["YTD Target"] > 0,
        out["YTD Achievement"] / out["YTD Target"],
        0.0,
    )
 
    out["Achievement Category"] = out["YTD Achievement %"].map(achievement_band)
    out["Current Run Rate"] = out["YTD Achievement"] / max(MONTHS_COMPLETED, 1)
    out["Estimated FY @ Current RR"] = out["Current Run Rate"] * 12.0
    out["Projected FY Achievement %"] = np.where(
        out["FY Target"] > 0,
        out["Estimated FY @ Current RR"] / out["FY Target"],
        0.0,
    )
 
    denominator = _num(final_target)
    if denominator is None or denominator <= 0:
        denominator = _num(out["FY Target"].sum()) or 0.0
 
    out["Contribution to Overall Target %"] = np.where(
        denominator > 0,
        out["Estimated FY @ Current RR"] / denominator,
        0.0,
    )
 
    # Asset-level achieved percentages are useful when drilling into an RM.
    for asset in ASSETS:
        fy = pd.to_numeric(subset[f"{sales}_{asset}_fy"], errors="coerce").fillna(0.0)
        ytd_target = pd.to_numeric(
            subset[f"{sales}_{asset}_ytd_tgt"], errors="coerce"
        ).fillna(0.0)
        ach = pd.to_numeric(subset[f"{sales}_{asset}_ach"], errors="coerce").fillna(0.0)
 
        out[f"{asset} YTD %"] = np.where(ytd_target > 0, ach / ytd_target, 0.0)
        out[f"{asset} Current RR"] = ach / max(MONTHS_COMPLETED, 1)
        out[f"{asset} Projected FY %"] = np.where(
            fy > 0,
            (ach / max(MONTHS_COMPLETED, 1) * 12.0) / fy,
            0.0,
        )
 
    out = out.sort_values(
        ["YTD Achievement %", "YTD Achievement"],
        ascending=[False, False],
    ).reset_index(drop=True)
 
    return out
 
 
def build_category_contribution(
    detail: pd.DataFrame,
    final_target: Optional[float] = None,
) -> pd.DataFrame:
    """
    Aggregate RM bands and quantify how many percentage points each band is
    projected to contribute to the FINAL FY target at the current run rate.
    """
    if detail.empty:
        return pd.DataFrame()
 
    target_denominator = _num(final_target)
    if target_denominator is None or target_denominator <= 0:
        target_denominator = _num(detail["FY Target"].sum()) or 0.0
 
    total_projected = _num(detail["Estimated FY @ Current RR"].sum()) or 0.0
 
    grouped = (
        detail.groupby("Achievement Category", dropna=False)
        .agg(
            **{
                "RM Count": ("Employee Name", "count"),
                "FY Target": ("FY Target", "sum"),
                "YTD Target": ("YTD Target", "sum"),
                "YTD Achievement": ("YTD Achievement", "sum"),
                "Current Run Rate": ("Current Run Rate", "sum"),
                "Estimated FY @ Current RR": ("Estimated FY @ Current RR", "sum"),
            }
        )
        .reindex(ACHIEVEMENT_BAND_ORDER, fill_value=0)
        .reset_index()
    )
 
    grouped["Current YTD Achievement %"] = np.where(
        grouped["YTD Target"] > 0,
        grouped["YTD Achievement"] / grouped["YTD Target"],
        0.0,
    )
    grouped["Category Projected FY %"] = np.where(
        grouped["FY Target"] > 0,
        grouped["Estimated FY @ Current RR"] / grouped["FY Target"],
        0.0,
    )
    grouped["Contribution to Overall Target %"] = np.where(
        target_denominator > 0,
        grouped["Estimated FY @ Current RR"] / target_denominator,
        0.0,
    )
    grouped["Share of Projected Sales %"] = np.where(
        total_projected != 0,
        grouped["Estimated FY @ Current RR"] / total_projected,
        0.0,
    )
 
    return grouped
 
 
def _final_vertical_target(
    final_metrics: Dict[str, Any],
    sales: str,
    vertical: str,
) -> Optional[float]:
    frame = final_metrics.get(sales)
    if not isinstance(frame, pd.DataFrame) or frame.empty or vertical not in frame.index:
        return None
    return _num(frame.loc[vertical].get("FY27 Target"))
 
 
def _final_vertical_ytd(
    final_metrics: Dict[str, Any],
    sales: str,
    vertical: str,
) -> Optional[float]:
    frame = final_metrics.get(sales)
    if not isinstance(frame, pd.DataFrame) or frame.empty or vertical not in frame.index:
        return None
    return _num(frame.loc[vertical].get("YTD"))
 
 
def _gold_star_card(rank: int, row: pd.Series) -> str:
    name = escape(str(row.get("Employee Name", "RM")))
    achievement = fmt_pct(row.get("YTD Achievement %"))
    projected = fmt_pct(row.get("Projected FY Achievement %"))
    contribution = fmt_pct(row.get("Contribution to Overall Target %"))
    rr = fmt_cr(row.get("Current Run Rate"))
    return (
        "<div class='gold-star-card'>"
        f"<div class='gold-star-rank'>Star #{rank}</div>"
        f"<div class='gold-star-name'>{name}</div>"
        "<div class='gold-star-stat'>"
        f"YTD achievement: <b>{achievement}</b><br>"
        f"Current run rate: <b>{rr}</b><br>"
        f"Projected FY: <b>{projected}</b><br>"
        f"Overall-target contribution: <b>{contribution}</b>"
        "</div></div>"
    )
 
 
def render_stars_of_month(detail: pd.DataFrame, vertical: str, sales: str) -> None:
    section("Stars of the Month")
 
    if detail.empty:
        st.info("No RM data is available for the selected view.")
        return
 
    # Workbook provides YTD/current-run-rate fields, not a standalone single-month
    # achievement column. Rank by YTD achievement % and then YTD achievement amount.
    ranked = detail.sort_values(
        ["YTD Achievement %", "YTD Achievement"],
        ascending=[False, False],
    ).reset_index(drop=True)
 
    top = ranked.head(3)
    columns = st.columns(max(len(top), 1))
    for index, (_, row) in enumerate(top.iterrows(), start=1):
        with columns[index - 1]:
            st.markdown(_gold_star_card(index, row), unsafe_allow_html=True)
 
    st.markdown(
        "<div class='note'>Stars are ranked from the performance fields available in the "
        "uploaded scorecard: overall YTD achievement versus YTD target, with YTD achievement "
        "amount as the tie-breaker. A standalone monthly actual field is not available, so "
        "the app does not invent a separate monthly score.</div>",
        unsafe_allow_html=True,
    )
 
    top_table = ranked.head(10).copy()
    columns_to_show = [
        column
        for column in [
            "Employee Name", "Emp Code", "ADID", "REGION", "EM City",
            "Achievement Category", "YTD Achievement %", "Current Run Rate",
            "Projected FY Achievement %", "Contribution to Overall Target %",
        ]
        if column in top_table.columns
    ]
    show_table(
        top_table[columns_to_show],
        {
            "Employee Name": "txt",
            "Emp Code": "txt",
            "ADID": "txt",
            "REGION": "txt",
            "EM City": "txt",
            "Achievement Category": "txt",
            "YTD Achievement %": "pct",
            "Current Run Rate": "cr",
            "Projected FY Achievement %": "pct",
            "Contribution to Overall Target %": "pct",
        },
    )
 
 
def render_rm_sales_segmentation(
    records: pd.DataFrame,
    final_metrics: Dict[str, Any],
    vertical: str,
    sales: str,
) -> None:
    final_target = _final_vertical_target(final_metrics, sales, vertical)
    final_ytd = _final_vertical_ytd(final_metrics, sales, vertical)
 
    detail = build_rm_performance_detail(
        records,
        vertical=vertical,
        sales=sales,
        final_target=final_target,
    )
    if detail.empty:
        st.info(f"No RM records are available for {vertical} · {SALES_LABEL[sales]}.")
        return
 
    contribution = build_category_contribution(detail, final_target=final_target)
 
    projected_total = _num(detail["Estimated FY @ Current RR"].sum()) or 0.0
    ytd_total = _num(detail["YTD Achievement"].sum()) or 0.0
    ytd_target_total = _num(detail["YTD Target"].sum()) or 0.0
    rm_count = len(detail)
 
    ytd_pct = ytd_total / ytd_target_total if ytd_target_total > 0 else 0.0
    projected_pct = (
        projected_total / final_target
        if final_target is not None and final_target > 0
        else (
            projected_total / (_num(detail["FY Target"].sum()) or 1.0)
        )
    )
 
    high_performers = int(
        detail["Achievement Category"]
        .isin(["100% and above", "90% - 100%"])
        .sum()
    )
 
    kpi_row([
        ("Total RMs", fmt_num(rm_count), f"{vertical} · {SALES_LABEL[sales]}", "off"),
        ("YTD Achievement vs YTD Target", fmt_pct(ytd_pct),
         fmt_cr(ytd_total), "off"),
        ("Projected FY Achievement", fmt_pct(projected_pct),
         fmt_cr(projected_total), "off"),
        ("90%+ RMs", fmt_num(high_performers),
         fmt_pct(high_performers / rm_count if rm_count else 0.0), "off"),
        ("FINAL FY27 Target", fmt_cr(final_target),
         "management target", "off"),
    ])
 
    section("RM Achievement Segmentation")
 
    count_map = dict(
        zip(
            contribution["Achievement Category"],
            contribution["RM Count"],
        )
    )
    count_cards = []
    for band in ACHIEVEMENT_BAND_ORDER:
        count_cards.append(
            (
                band,
                fmt_num(count_map.get(band, 0)),
                "RMs",
                "off",
            )
        )
    # Keep six bands readable by using two rows.
    kpi_row(count_cards[:3])
    kpi_row(count_cards[3:])
 
    count_chart = (
        contribution[["Achievement Category", "RM Count"]]
        .set_index("Achievement Category")
        .reindex(ACHIEVEMENT_BAND_ORDER)
    )
    st.bar_chart(count_chart)
 
    section("Category Run-Rate Contribution to Overall Target")
    st.markdown(
        "<div class='note'>Each category's RMs are annualised at their current run rate. "
        "The projected amount is then divided by the FY27 target from the FINAL sheet for "
        "the same channel and sales basis. Therefore the contribution percentages show "
        "how many percentage points each RM category is expected to add to the overall "
        "target achievement if its current pace continues.</div>",
        unsafe_allow_html=True,
    )
 
    contribution_display = contribution.copy()
    show_table(
        contribution_display,
        {
            "Achievement Category": "txt",
            "RM Count": "num",
            "FY Target": "cr",
            "YTD Target": "cr",
            "YTD Achievement": "cr",
            "Current YTD Achievement %": "pct",
            "Current Run Rate": "cr",
            "Estimated FY @ Current RR": "cr",
            "Category Projected FY %": "pct",
            "Contribution to Overall Target %": "pct",
            "Share of Projected Sales %": "pct",
        },
    )
 
    contribution_chart = contribution[
        ["Achievement Category", "Contribution to Overall Target %"]
    ].copy()
    contribution_chart["Contribution to Overall Target %"] *= 100.0
    contribution_chart = contribution_chart.set_index("Achievement Category").reindex(
        ACHIEVEMENT_BAND_ORDER
    )
    st.bar_chart(contribution_chart)
 
    total_contribution = _num(
        contribution["Contribution to Overall Target %"].sum()
    ) or 0.0
    st.markdown(
        "<div class='final-output-box'>"
        f"<b>Final run-rate output · {vertical} · {SALES_LABEL[sales]}</b><br>"
        f"At the current RM run rates, the six categories together are projected to deliver "
        f"<b>{fmt_pct(total_contribution)}</b> of the FINAL FY27 target"
        + (
            f" ({fmt_cr(final_target)})."
            if final_target is not None
            else "."
        )
        + "</div>",
        unsafe_allow_html=True,
    )
 
    section("RM Drill-Down by Achievement Category")
    selected_band = st.selectbox(
        "Achievement Category",
        ACHIEVEMENT_BAND_ORDER,
        index=0,
        key=f"rm_band_{vertical}_{sales}",
    )
 
    rm_rows = detail.loc[detail["Achievement Category"] == selected_band].copy()
    if rm_rows.empty:
        st.info(f"No RMs fall in {selected_band} for this selection.")
    else:
        display_columns = [
            column
            for column in [
                "Employee Name", "Emp Code", "ADID", "ZONE", "REGION", "EM City",
                "Achievement Category", "FY Target", "YTD Target", "YTD Achievement",
                "YTD Achievement %", "Current Run Rate", "Estimated FY @ Current RR",
                "Projected FY Achievement %", "Contribution to Overall Target %",
                "Equity YTD %", "Debt YTD %", "Liquid YTD %",
            ]
            if column in rm_rows.columns
        ]
        show_table(
            rm_rows[display_columns],
            {
                "Employee Name": "txt",
                "Emp Code": "txt",
                "ADID": "txt",
                "ZONE": "txt",
                "REGION": "txt",
                "EM City": "txt",
                "Achievement Category": "txt",
                "FY Target": "cr",
                "YTD Target": "cr",
                "YTD Achievement": "cr",
                "YTD Achievement %": "pct",
                "Current Run Rate": "cr",
                "Estimated FY @ Current RR": "cr",
                "Projected FY Achievement %": "pct",
                "Contribution to Overall Target %": "pct",
                "Equity YTD %": "pct",
                "Debt YTD %": "pct",
                "Liquid YTD %": "pct",
            },
        )
 
    render_stars_of_month(detail, vertical, sales)
 
 
def make_rm_segmentation_export(
    records: pd.DataFrame,
    final_metrics: Dict[str, Any],
) -> bytes:
    """Downloadable workbook covering every vertical and both sales bases."""
    output = io.BytesIO()
 
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for vertical in VERTICALS:
            for sales in SALES_TYPES:
                final_target = _final_vertical_target(final_metrics, sales, vertical)
                detail = build_rm_performance_detail(
                    records, vertical, sales, final_target
                )
                contribution = build_category_contribution(detail, final_target)
                stars = (
                    detail.sort_values(
                        ["YTD Achievement %", "YTD Achievement"],
                        ascending=[False, False],
                    )
                    .head(10)
                    .copy()
                )
 
                prefix = f"{vertical}-{sales}"
                detail.to_excel(
                    writer,
                    sheet_name=f"{prefix}-RM"[:31],
                    index=False,
                )
                contribution.to_excel(
                    writer,
                    sheet_name=f"{prefix}-Bands"[:31],
                    index=False,
                )
                stars.to_excel(
                    writer,
                    sheet_name=f"{prefix}-Stars"[:31],
                    index=False,
                )
 
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            for cells in ws.columns:
                width = max(
                    (len(str(cell.value or "")) for cell in cells[:150]),
                    default=10,
                )
                ws.column_dimensions[cells[0].column_letter].width = min(
                    max(width + 2, 12),
                    36,
                )
 
    output.seek(0)
    return output.getvalue()
 
 
 
def _clean_filter_values(series: pd.Series) -> List[str]:
    """Return sorted non-empty values for RM filter dropdowns."""
    cleaned = (
        series.astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
    )
    cleaned = cleaned[
        cleaned.ne("")
        & cleaned.str.lower().ne("nan")
        & cleaned.str.lower().ne("none")
    ]
    return sorted(cleaned.drop_duplicates().tolist(), key=lambda value: value.lower())
 
 
def _apply_exact_text_filter(
    frame: pd.DataFrame,
    column: str,
    selected: str,
) -> pd.DataFrame:
    """Apply one exact dropdown filter while treating 'All' as no filter."""
    if selected == "All" or column not in frame.columns:
        return frame
 
    values = (
        frame[column]
        .astype(str)
        .str.replace("\u00a0", " ", regex=False)
        .str.strip()
    )
    return frame.loc[values == selected].copy()
 
 
def render_retail_rm_filters(records: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Cascading Retail-only filters for the second page.
 
    Filter order:
      ZONE -> REGION -> MKT TYPE
 
    The returned dataframe keeps DHNI/VRM untouched and replaces only Retail
    rows with the filtered Retail population. This means the same dataframe can
    safely be used by the page and by the downloadable workbook.
    """
    retail = records.loc[records["Vertical"] == "Retail"].copy()
 
    selections: Dict[str, str] = {
        "ZONE": "All",
        "REGION": "All",
        "MKT TYPE": "All",
    }
 
    st.markdown(
        "<div class='subsection-title'>Retail RM Filters</div>",
        unsafe_allow_html=True,
    )
 
    filter_cols = st.columns(3)
 
    # -------------------------
    # ZONE
    # -------------------------
    zone_options = ["All"]
    if "ZONE" in retail.columns:
        zone_options += _clean_filter_values(retail["ZONE"])
 
    with filter_cols[0]:
        selections["ZONE"] = st.selectbox(
            "ZONE",
            zone_options,
            index=0,
            key="retail_rm_zone_filter",
            help="Filter Retail RMs by zone.",
        )
 
    filtered_retail = _apply_exact_text_filter(
        retail,
        "ZONE",
        selections["ZONE"],
    )
 
    # -------------------------
    # REGION — cascades from ZONE
    # -------------------------
    region_options = ["All"]
    if "REGION" in filtered_retail.columns:
        region_options += _clean_filter_values(filtered_retail["REGION"])
 
    with filter_cols[1]:
        selections["REGION"] = st.selectbox(
            "REGION",
            region_options,
            index=0,
            key=f"retail_rm_region_filter_{selections['ZONE']}",
            help="Region options automatically narrow after selecting a zone.",
        )
 
    filtered_retail = _apply_exact_text_filter(
        filtered_retail,
        "REGION",
        selections["REGION"],
    )
 
    # -------------------------
    # MKT TYPE — cascades from ZONE + REGION
    # -------------------------
    market_options = ["All"]
    if "MKT TYPE" in filtered_retail.columns:
        market_options += _clean_filter_values(filtered_retail["MKT TYPE"])
 
    with filter_cols[2]:
        selections["MKT TYPE"] = st.selectbox(
            "MKT TYPE",
            market_options,
            index=0,
            key=(
                f"retail_rm_market_filter_"
                f"{selections['ZONE']}_{selections['REGION']}"
            ),
            help="Market Type options narrow after the selected zone and region.",
        )
 
    filtered_retail = _apply_exact_text_filter(
        filtered_retail,
        "MKT TYPE",
        selections["MKT TYPE"],
    )
 
    # Keep all non-Retail rows unchanged. Replace Retail with the selected slice.
    non_retail = records.loc[records["Vertical"] != "Retail"].copy()
    filtered_records = pd.concat(
        [filtered_retail, non_retail],
        ignore_index=True,
        sort=False,
    )
 
    active_filters = [
        f"{column}: {value}"
        for column, value in selections.items()
        if value != "All"
    ]
 
    active_text = " · ".join(active_filters) if active_filters else "All Retail RMs"
 
    st.markdown(
        "<div class='callout'>"
        f"<b>Retail population:</b> {len(filtered_retail):,} RM(s)<br>"
        f"<b>Active filters:</b> {escape(active_text)}"
        "</div>",
        unsafe_allow_html=True,
    )
 
    return filtered_records, selections
 
 
 
def render_rm_segmentation_page(records: pd.DataFrame, payload: bytes) -> None:
    """Second application page dedicated to RM achievement segmentation."""
    final_metrics = parse_final_dashboard_metrics(payload)
 
    st.markdown(
        "<div class='app-title'>RM Performance Segmentation & Contribution Analysis</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='app-sub'>Retail · DHNI · VRM | Achievement bands → run-rate "
        "contribution to FINAL target → Stars of the Month</div>",
        unsafe_allow_html=True,
    )
    st.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
 
    st.markdown(
        "<span class='category-pill'>100% and above</span>"
        "<span class='category-pill'>90% - 100%</span>"
        "<span class='category-pill'>75% - 90%</span>"
        "<span class='category-pill'>50% - 75%</span>"
        "<span class='category-pill'>30% - 50%</span>"
        "<span class='category-pill'>Less than 30%</span>",
        unsafe_allow_html=True,
    )
 
    vertical = st.selectbox(
        "Select RM Channel",
        VERTICALS,
        index=0,
        key="rm_seg_vertical",
        help="Switch between Retail, DHNI and VRM.",
    )
 
    page_records = records
    retail_filter_selections: Dict[str, str] = {
        "ZONE": "All",
        "REGION": "All",
        "MKT TYPE": "All",
    }
 
    # Retail gets the requested extra management filters.
    # DHNI and VRM continue to use their full populations.
    if vertical == "Retail":
        page_records, retail_filter_selections = render_retail_rm_filters(records)
 
    tabs = st.tabs([SALES_LABEL["GS"], SALES_LABEL["NS"]])
    for tab, sales in zip(tabs, SALES_TYPES):
        with tab:
            render_rm_sales_segmentation(
                page_records,
                final_metrics,
                vertical,
                sales,
            )
 
    section("Download RM Segmentation Analysis")
    try:
        export_payload = make_rm_segmentation_export(page_records, final_metrics)
        st.download_button(
            "Download Full RM Segmentation Workbook",
            data=export_payload,
            file_name="rm_performance_segmentation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if vertical == "Retail":
            active = [
                f"{name}: {value}"
                for name, value in retail_filter_selections.items()
                if value != "All"
            ]
            st.caption(
                "Retail sheets in this download use the current filters"
                + (f" ({' · '.join(active)})." if active else " (All Retail RMs).")
                + " DHNI and VRM remain unfiltered."
            )
    except Exception:
        st.warning("The RM segmentation export could not be generated.")
 
    st.markdown(
        "<div class='note'>Missing / undefined numeric outputs are displayed as 0, "
        "as requested.</div>",
        unsafe_allow_html=True,
    )
 
 
def render_page_navigation() -> str:
    st.sidebar.markdown(
        "<div class='sidebar-title'>Application Page</div>",
        unsafe_allow_html=True,
    )
    page = st.sidebar.selectbox(
        "Application Page",
        [
            "Executive Scenario Dashboard",
            "RM Performance Segmentation",
        ],
        index=0,
        key="application_page_selector",
        label_visibility="collapsed",
    )
    st.sidebar.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
    return page
 
 
 
# =============================================================================
# 16. APPLICATION ENTRY POINT
# =============================================================================
 
def render_channel_simulator(model: ScenarioModel, basis: str) -> None:
    section("Scenario 8 · Channel Growth & Target Simulator")
    frame, formats = build_channel_scenario_analysis(model, basis)
    if frame.empty:
        callout("No mapped channel data is available. Use Channel mapping in the sidebar to classify the workbook.", tone="warn")
        return
    st.markdown("<div class='channel-simulator-note'>Nine independent management levers · MoM growth · January 2027 target · March 2027 target · 20% default leakage</div>", unsafe_allow_html=True)
    show_table(frame, formats)
    jan_gap = frame["Jan Gap / Headroom"].sum()
    mar_gap = frame["Mar Gap / Headroom"].sum()
    total_incremental = frame["March Incremental Sales"].sum()
    kpi_row([
        ("January Portfolio Headroom", fmt_cr_signed(jan_gap), "positive = above target"),
        ("March Portfolio Headroom", fmt_cr_signed(mar_gap), "positive = above target"),
        ("Incremental March Sales", fmt_cr_signed(total_incremental), "vs current projection"),
        ("Channels", f"{len(frame)} / {len(CHANNELS)}", "mapped into simulator", "off"),
    ])
    if (frame["Jan Gap / Headroom"] >= 0).all() and (frame["Mar Gap / Headroom"] >= 0).all():
        callout("<span class='tag-ok'>✓ ALL CHANNELS ON TRACK</span> — selected growth assumptions clear both January and March targets after leakage.", tone="ok")
    else:
        misses = frame.loc[(frame["Jan Gap / Headroom"] < 0) | (frame["Mar Gap / Headroom"] < 0), "Channel"].tolist()
        callout("<span class='tag-warn'>⚠ CHANNEL GAP</span> — review: " + ", ".join(misses) + ". Increase MoM growth or adjust the relevant target.", tone="warn")
 
 
def render_channel_optimizer(model: ScenarioModel, basis: str) -> None:
    section("Scenario 9 · Channel Mix Optimiser")
    target = float(model.params.get("optimizer_target", 1.20))
    frame, formats = build_channel_scenario_analysis(model, basis)
    if frame.empty:
        st.info("No mapped channel data is available.")
        return
    # Scenario 9 grid already solves each channel against the target controls; expose the
    # optimiser as a management view and keep the target visible as the portfolio ambition.
    frame = frame.copy()
    frame["Optimised MoM"] = frame["MoM Growth"]
    frame["Portfolio Ambition"] = target
    formats = dict(formats); formats.update({"Optimised MoM":"pct_signed", "Portfolio Ambition":"pct"})
    show_table(frame, formats)
    weighted = model.cell(basis)
    callout(f"The optimiser is solving the minimum momentum trajectory by channel while protecting the selected portfolio March ambition of {fmt_pct(target)} and the January milestone. Adjust channel mappings and targets in the sidebar to change the optimisation universe.")
 
 
def render_dashboard(records: pd.DataFrame, payload: bytes) -> None:
    scenario_id, params, basis, mapping = render_sidebar(records)
    records = map_business_segments(records, mapping)
    channel_mapping = params.get("channel_mapping") or st.session_state.get("channel_mapping") or {}
    records = map_business_channels(records, channel_mapping)
    segment_counts = segment_diagnostics(records)
    grid = build_base_grid(records)
    model = ScenarioModel(scenario_id, grid, params)
 
    final_metrics = parse_final_dashboard_metrics(payload)
 
    st.markdown(f"<div class='app-title'>{APP_TITLE}</div>", unsafe_allow_html=True)
    st.markdown(
        "<div class='app-sub'>FINAL management metrics → selected scenario comparison → "
        "required future run rate → revenue impact</div>",
        unsafe_allow_html=True,
    )
    st.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
 
    # Single-screen management flow:
    # 1. FINAL current metrics
    # 2. selected scenario against the same metrics
    # 3. all existing scenario / revenue detail
    render_final_metric_baseline(final_metrics, model)
    render_final_scenario_comparison(final_metrics, model, basis)
    render_current_runrate_metric_grid(final_metrics, model)
 
    bundle = render_revenue_kpis(model, basis)
    render_detail_expander(model)
 
    render_vertical_section(model)
    render_asset_section(model)
 
    if scenario_id == 6:
        render_segment_section(model, basis, segment_counts)
    if scenario_id == 7:
        render_momentum_section(model, basis)
    if scenario_id == 8:
        render_channel_simulator(model, basis)
    if scenario_id == 9:
        render_channel_optimizer(model, basis)
 
    render_revenue_detail(model, basis, bundle)
 
    with st.expander("Source FINAL sheet", expanded=False):
        st.markdown(
            "<div class='note'>Reference view of the workbook's FINAL sheet. "
            "It is part of this same dashboard screen and is not a separate application view.</div>",
            unsafe_allow_html=True,
        )
        st.markdown(build_final_sheet_html(payload), unsafe_allow_html=True)
 
    render_export(model, basis)
 
 
def reset_workbook() -> None:
    for key in ("workbook", "segment_mapping", "channel_mapping", "application_page_selector"):
        st.session_state.pop(key, None)
    rerun()
 
 
def main() -> None:
    st.set_page_config(
        page_title=APP_TITLE, page_icon="▮", layout="wide", initial_sidebar_state="expanded",
    )
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
 
    if "workbook" not in st.session_state:
        render_upload_screen()
        return
 
    try:
        records = load_workbook(st.session_state["workbook"])
    except WorkbookError as error:
        st.error(str(error))
        if st.button("Use another workbook"):
            reset_workbook()
        return
    except Exception:  # never surface a raw traceback to management
        st.error(
            "The workbook could not be read. Please upload the standard RM scorecard workbook "
            "containing RM Retail Sales, RM DHNI, VRM and FINAL."
        )
        if st.button("Use another workbook"):
            reset_workbook()
        return
 
    page = render_page_navigation()
 
    try:
        if page == "RM Performance Segmentation":
            render_rm_segmentation_page(records, st.session_state["workbook"])
            st.sidebar.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)
            if st.sidebar.button("Use another workbook", key="rm_page_change_workbook"):
                reset_workbook()
        else:
            render_dashboard(records, st.session_state["workbook"])
    except Exception:  # never surface a raw traceback to management
        st.error(
            "This view could not be prepared from the uploaded workbook. Please select another "
            "page/scenario, or upload a workbook that matches the standard RM scorecard format."
        )
 
 
# =============================================================================
# 17. MANAGEMENT CUTS + ALL-SCENARIO VIEW (FINAL-aware)
# =============================================================================

LOCATION_FIELD_ALIASES = ["MKT TYPE", "MKT TYPE ", "Market Type", "Mkt Type"]

def _location_column(records: pd.DataFrame) -> Optional[str]:
    for c in LOCATION_FIELD_ALIASES:
        if c in records.columns:
            return c
    return None

def _location_options(records: pd.DataFrame, channel: str) -> List[str]:
    col = _location_column(records)
    if col is None:
        return ["All"]
    work = records
    if channel != "All" and "Vertical" in work.columns:
        work = work.loc[work["Vertical"] == channel]
    vals = sorted({str(v).strip() for v in work[col].dropna().tolist() if str(v).strip()})
    # Preserve the workbook's location/mkt-type cuts, including T2/T6/T30/B30/EM.
    return ["All"] + vals

def _apply_management_cuts(
    records: pd.DataFrame,
    channel: str,
    location: str,
) -> pd.DataFrame:
    out = records.copy()
    if channel != "All" and "Vertical" in out.columns:
        out = out.loc[out["Vertical"] == channel].copy()
    col = _location_column(out)
    if location != "All" and col is not None:
        out = out.loc[out[col].astype(str).str.strip() == location].copy()
    return out

def render_management_cut_controls(records: pd.DataFrame) -> Tuple[str, str, str, str, pd.DataFrame]:
    """
    Four explicit management cuts:
      1) Gross Sales / Net Sales
      2) Channel: Retail / DHNI / VRM
      3) Asset Class: Equity / Debt / Liquid
      4) Location / Market Type: T2 / T6 / T30 / B30 / EM / workbook values

    The first three are visualization cuts. Channel + location are applied to
    the calculation grid so scenario mathematics is recalculated on the exact
    selected population.
    """
    section("Management Cuts")
    st.markdown(
        "<div class='note'>All scenario calculations below are recalculated on the "
        "selected Channel + Location population. Gross/Net Sales and Asset Class "
        "control the visualised metric without changing the underlying workbook data.</div>",
        unsafe_allow_html=True,
    )
    c1, c2, c3, c4 = st.columns(4)

    with c1:
        sales = st.selectbox(
            "Sales",
            ["Gross Sales", "Net Sales"],
            index=0,
            key="mgmt_sales_cut",
        )
    with c2:
        channel_options = ["All"] + [v for v in VERTICALS if v in set(records.get("Vertical", pd.Series(dtype=str)).astype(str))]
        channel = st.selectbox(
            "Channel",
            channel_options,
            index=0,
            key="mgmt_channel_cut",
        )
    with c3:
        asset = st.selectbox(
            "Asset Class",
            ["All", *ASSETS],
            index=0,
            key="mgmt_asset_cut",
        )
    with c4:
        loc_options = _location_options(records, channel)
        current_loc = st.session_state.get("mgmt_location_cut", "All")
        loc_index = loc_options.index(current_loc) if current_loc in loc_options else 0
        location = st.selectbox(
            "Location / Market Type",
            loc_options,
            index=loc_index,
            key="mgmt_location_cut",
        )

    filtered = _apply_management_cuts(records, channel, location)
    if filtered.empty:
        st.warning("No employee records match the selected Channel + Location cut.")
    return sales, channel, asset, location, filtered

def _scenario_default_params(scenario_id: int) -> Dict[str, Any]:
    """Defaults exactly from the existing scenario engine configuration."""
    return {
        "dip": S3_DEFAULT_DIP,
        "jan_target": S7_DEFAULT_JAN_TARGET,
        "mar_target": S7_DEFAULT_MAR_TARGET,
        "leakage": S7_DEFAULT_LEAKAGE,
        "channel_growth": dict(S8_DEFAULT_GROWTH),
        "channel_jan_target": dict(S8_DEFAULT_JAN_TARGET),
        "channel_mar_target": dict(S8_DEFAULT_MAR_TARGET),
        "optimizer_target": 1.20,
        "channel_mapping": {},
    }

def _selected_metric_cell(
    model: "ScenarioModel",
    sales: str,
    asset: str,
) -> Dict[str, Any]:
    sales_key = "GS" if sales == "Gross Sales" else "NS"
    return model.cell(
        sales_key,
        asset=None if asset == "All" else asset,
    )

def _render_cut_summary(
    model: "ScenarioModel",
    sales: str,
    asset: str,
    channel: str,
    location: str,
) -> None:
    sales_key = "GS" if sales == "Gross Sales" else "NS"
    cell = _selected_metric_cell(model, sales, asset)

    section(f"{sales} · Current State")
    active = [x for x in [channel if channel != "All" else None,
                          asset if asset != "All" else None,
                          location if location != "All" else None] if x]
    scope = " · ".join(active) if active else "All business"
    st.markdown(f"<div class='note'><b>Scope:</b> {escape(scope)}</div>", unsafe_allow_html=True)

    kpi_row([
        ("FY Target", fmt_cr(cell.get("fy_target")), "Workbook target"),
        ("YTD June Target", fmt_cr(cell.get("ytd_target")), "Workbook YTD target"),
        ("YTD Achievement", fmt_cr(cell.get("ytd_ach")), 
         fmt_pct(cell.get("ytd_ach_pct")) if _num(cell.get("ytd_ach_pct")) is not None else None),
        ("Current Monthly RR", fmt_cr(cell.get("current_rr")), "YTD ÷ 3"),
        ("Current March Projection", fmt_cr(cell.get("current_march")),
         fmt_pct(cell.get("current_march_pct")) if _num(cell.get("current_march_pct")) is not None else None),
    ])

    # Asset-class cut is intentionally visible even when Asset Class = All.
    if asset == "All":
        rows = []
        for a in ASSETS:
            c = model.cell(sales_key, asset=a)
            rows.append({
                "Asset Class": a,
                "FY Target": c.get("fy_target"),
                "YTD Target": c.get("ytd_target"),
                "YTD Achievement": c.get("ytd_ach"),
                "Achievement %": c.get("ytd_ach_pct"),
                "Current RR": c.get("current_rr"),
                "Current March Projection": c.get("current_march"),
                "Projected FY %": c.get("current_march_pct"),
            })
        frame = pd.DataFrame(rows)
        show_table(frame, {
            "Asset Class": "txt", "FY Target": "cr", "YTD Target": "cr",
            "YTD Achievement": "cr", "Achievement %": "pct", "Current RR": "cr",
            "Current March Projection": "cr", "Projected FY %": "pct",
        })

def _build_all_scenario_matrix(
    filtered_grid: pd.DataFrame,
    selected_scenario_id: int,
    selected_params: Dict[str, Any],
    sales_key: str,
    asset: str,
) -> pd.DataFrame:
    """Calculate Scenarios 1-9 on the same selected cut."""
    rows: List[Dict[str, Any]] = []
    if filtered_grid.empty:
        return pd.DataFrame()

    for sid in SCENARIO_ORDER:
        params = _scenario_default_params(sid)
        if sid == selected_scenario_id:
            # Sidebar controls are authoritative for the selected scenario.
            params.update(selected_params)
        try:
            model = ScenarioModel(sid, filtered_grid, params)
            cell = model.cell(sales_key, asset=None if asset == "All" else asset)
            rows.append({
                "Scenario": f"Scenario {sid}",
                "Scenario Name": SCENARIOS[sid]["name"],
                "FY Target": cell.get("fy_target"),
                "Current YTD": cell.get("ytd_ach"),
                "Current March Projection": cell.get("current_march"),
                "Scenario March Estimate": cell.get("march_amount"),
                "Scenario March %": cell.get("march_pct"),
                "Run Rate": cell.get("scen_rr"),
                "Run Rate Change %": cell.get("rr_change_pct"),
                "Required / March Target": cell.get("march_required"),
                "Headroom / Gap": cell.get("headroom_amt"),
            })
        except Exception as exc:
            rows.append({
                "Scenario": f"Scenario {sid}",
                "Scenario Name": SCENARIOS[sid]["name"],
                "Error": str(exc),
            })
    return pd.DataFrame(rows)

def render_all_scenario_matrix(
    filtered_grid: pd.DataFrame,
    selected_scenario_id: int,
    selected_params: Dict[str, Any],
    sales: str,
    asset: str,
) -> None:
    sales_key = "GS" if sales == "Gross Sales" else "NS"
    section("All Scenarios · Selected Cut")
    frame = _build_all_scenario_matrix(
        filtered_grid, selected_scenario_id, selected_params, sales_key, asset
    )
    if frame.empty:
        st.info("No scenario output is available for this cut.")
        return

    formats = {
        "Scenario": "txt", "Scenario Name": "txt", "FY Target": "cr",
        "Current YTD": "cr", "Current March Projection": "cr",
        "Scenario March Estimate": "cr", "Scenario March %": "pct",
        "Run Rate": "cr", "Run Rate Change %": "pct_signed",
        "Required / March Target": "cr", "Headroom / Gap": "cr_signed",
    }
    show_table(frame, formats)

    selected = frame.loc[frame["Scenario"] == f"Scenario {selected_scenario_id}"]
    if not selected.empty:
        row = selected.iloc[0]
        st.markdown(
            f"<div class='scenario-highlight'><b>Selected: Scenario {selected_scenario_id} · "
            f"{SCENARIOS[selected_scenario_id]['name']}</b><br>"
            f"{SCENARIOS[selected_scenario_id]['explanation']}<br>"
            f"<span class='s-milestone'>{SCENARIOS[selected_scenario_id]['milestone']}</span></div>",
            unsafe_allow_html=True,
        )

def render_final_reference_with_cuts(payload: bytes) -> None:
    """
    Keep FINAL as the workbook reference surface. This intentionally shows
    the workbook values without rewriting them into model outputs.
    """
    with st.expander("FINAL sheet · source of truth", expanded=False):
        st.markdown(
            "<div class='note'>This is the uploaded workbook's FINAL sheet. "
            "Its current targets, YTD achievement, AUM, run-rate estimations, "
            "location cuts and scenario parameter blocks remain available for audit.</div>",
            unsafe_allow_html=True,
        )
        try:
            raw = load_final_sheet_frame(payload)
            st.dataframe(raw, use_container_width=True, hide_index=True, height=620)
        except Exception as exc:
            st.warning(f"FINAL sheet could not be displayed: {escape(str(exc))}")

def render_dashboard_with_management_cuts(records: pd.DataFrame, payload: bytes) -> None:
    """
    Replacement dashboard entry point:
      - FINAL is loaded and displayed as the source layer.
      - Management cuts are explicit dropdowns.
      - Channel + Location recalculate the analytical grid.
      - Asset + Sales select the displayed metric.
      - Scenarios 1-9 are all evaluated for the selected cut.
    """
    scenario_id, params, _basis, mapping = render_sidebar(records)

    records = map_business_segments(records, mapping)
    channel_mapping = params.get("channel_mapping") or st.session_state.get("channel_mapping") or {}
    records = map_business_channels(records, channel_mapping)

    final_metrics = parse_final_dashboard_metrics(payload)

    st.markdown(
        "<div class='app-title'>Sales Target Achievement · FINAL + Scenario Planner</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<div class='app-sub'>FINAL workbook values → management cuts → current state → "
        "all scenario projections → selected scenario detail</div>",
        unsafe_allow_html=True,
    )
    st.markdown("<div class='app-rule'></div>", unsafe_allow_html=True)

    sales, channel, asset, location, filtered_records = render_management_cut_controls(records)

    if filtered_records.empty:
        render_final_reference_with_cuts(payload)
        return

    grid = build_base_grid(filtered_records)
    model = ScenarioModel(scenario_id, grid, params)

    # Keep FINAL visible for the unfiltered management view, while the selected
    # cut always uses the same employee-level calculation engine.
    if channel == "All" and location == "All":
        render_final_metric_baseline(final_metrics, model)

    _render_cut_summary(model, sales, asset, channel, location)

    # Scenario 1-9 matrix is always calculated, not only the selected scenario.
    render_all_scenario_matrix(grid, scenario_id, params, sales, asset)

    # Detailed selected-scenario outputs continue to use the original engine.
    section(f"Scenario {scenario_id} · {SCENARIOS[scenario_id]['name']} · Detailed Output")
    if scenario_id == 6:
        segment_counts = segment_diagnostics(filtered_records)
        render_segment_section(model, "NS", segment_counts)
    elif scenario_id == 7:
        render_momentum_section(model, "NS")
    elif scenario_id == 8:
        render_channel_simulator(model, "NS")
    elif scenario_id == 9:
        render_channel_optimizer(model, "NS")
    else:
        # Existing detailed asset/vertical calculations, now on the selected cut.
        render_vertical_section(model)
        render_asset_section(model)

    render_final_reference_with_cuts(payload)
    render_export(model, "NS")

# Use the new FINAL-aware management dashboard without changing the existing
# scenario calculation engine.
_ORIGINAL_RENDER_DASHBOARD = render_dashboard
render_dashboard = render_dashboard_with_management_cuts

if __name__ == "__main__":
    main()
